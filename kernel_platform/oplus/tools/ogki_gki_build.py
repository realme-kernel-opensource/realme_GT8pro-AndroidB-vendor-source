import os
import argparse
import re
import sys
import subprocess
import shutil
from datetime import datetime
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import urllib3
import glob
import csv

file_paths = [
    ("boot.img", ""),
    ("boot-lz4.img", ""),
    ("boot-gz.img", ""),
    ("system_dlkm.flatten.erofs.img", "system_dlkm.img"),
    ("vmlinux", ""),
    ("gki-info.txt", ""),
    ("vmlinux.symvers", "")
]

repo_cmd = "repo"
# repo_cmd = "aosp/google-aosp-android15-6.6/.repo/repo/repo"
base_path = ""
dump_url = "https://android-review.googlesource.com/c/kernel/common/+/"


def get_tools_path(relative_path):
    global base_path

    if not base_path:
        base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    # print("base_path {} relative_path {} full_path {}".format(base_path, relative_path, full_path))
    return full_path


def extract_android_version(branch):
    pattern = r'android\d+-\d+(\.\d+)?'
    match = re.search(pattern, branch)
    if match:
        return match.group()
    else:
        return None


def get_current_date():
    now = datetime.now()
    return now.strftime('%Y-%m')


def is_valid_date_format(date):
    try:
        if date is None:
            return False
        date_pattern = re.compile(r'^\d{4}-\d{2}$')
        return bool(date_pattern.match(date))
    except (TypeError, re.error):
        return False


def generate_build_file_name(args, branch, date, release, number):
    date_prefix = f"{branch}-{date}_" if is_valid_date_format(date) else f"{branch}_"
    release_part = f"r{release}"
    oem_suffix = f"_ab{args.oem_ki.lower()}{number}" if args.oem_ki.lower() == "ogki" else ""

    file_name = f"{date_prefix}{release_part}{oem_suffix}"
    print(f"Generated file name: {file_name}")

    return file_name


def extract_and_increment(text):
    pattern = r'r(\d+)'
    match = re.search(pattern, text)

    if match:
        result = int(match.group(1)) + 1
        # result = int(match.group(1))
        return result
    else:
        return None


def process_config_file(config):
    if not os.path.exists(config):
        print(f"The file {config} does not exist.")
        return 1

    patterns = [
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+\b"),
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+_abogki\d+\b"),
        re.compile(r"android\d+-\d+\.\d+_r\d+_abogki\d+\b"),
        re.compile(r"android\d+-\d+\.\d+_r\d+\b")
    ]
    results = []

    try:
        with open(config, "r") as file:
            for line in file:
                line = line.strip()  # 去除前后空白字符
                for pattern in patterns:
                    match = pattern.search(line)
                    # print('pattern:', pattern)
                    # print('line:', line)
                    # print('match:', match)
                    if match:
                        result = extract_and_increment(match.group())
                        results.append(result)
    except Exception as e:
        print(f"Error processing the file {config}: {e}")
        return 1

    if results:
        return max(results)
    else:
        return 1


def get_current_release(args, out):
    command = [
        "python3", get_tools_path("ogki_gki_artifactory.py"),
        "-t", "get",
        "-k", args.oem_ki.upper(),
        "-c", "release.txt",
        "-i", out
    ]
    command_string = subprocess.list2cmdline(command)
    print(command_string)
    print('\ncommand', command)
    print('\n')

    try:
        result = subprocess.run(command, shell=False, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except subprocess.CalledProcessError as e:
        print("run command error： {}".format(e))
    release = process_config_file(os.path.join(out, "release.txt"))
    print('\ncurrent_release', release)
    # sys.exit(1)
    return release


def check_if_partner(input_str):
    if "partner-android" in input_str:
        return "partner", "ogki"
    else:
        return "aosp", "gki"


def process_path(path, ack_type, branch):
    # 去除路径最后的'/'
    path = path.rstrip('/')
    # 获取路径的最后一个目录名
    last_dir = os.path.basename(path)
    parent_dir = os.path.dirname(path)

    # 检查并追加type变量的值
    if ack_type not in last_dir:
        last_dir += "-{}".format(ack_type)

    # 检查并追加branch变量的值
    if branch not in last_dir:
        last_dir += "-{}".format(branch)

    # 生成新路径
    new_path = os.path.join(parent_dir, last_dir)
    return new_path


def create_symlink(src, dst):
    if not os.path.exists(src):
        print("Source directory does not exist:", src)
        return
    if os.path.exists(dst):
        print("Destination directory already exists:", dst)
        return

    # Calculate the relative path from dst to src
    src_dir = os.path.dirname(src)
    dst_dir = os.path.dirname(dst)
    relative_src = os.path.relpath(src_dir, dst_dir)

    # Create the symlink with the relative path
    os.symlink(os.path.join(relative_src, os.path.basename(src)), dst)
    print("Soft link created successfully:", dst, "->", os.path.join(relative_src, os.path.basename(src)))


def delete_symlink(path):
    if os.path.islink(path):
        os.unlink(path)
        print("Symbolic link deleted successfully:", path)
    else:
        print("Path is not a symbolic link:", path)


def is_token_valid(token):
    """
    验证 token 是否为一个有效的数字串
    不能有/ 或者 其他字符
    """
    return token.isdigit()


def parse_gerrit_number(input_str):
    """
    解析输入字符串，提取最后两个数字并返回特定格式化字符串，以及状态。

    Args:
        input_str (str): 输入的字符串。

    Returns:
        tuple: 格式化字符串和状态（True/False）。
    """
    # 使用正则表达式找到所有数字
    all_tokens = re.split(r'/+', input_str)

    # 筛选出有效的数字token
    valid_numbers = [token for token in all_tokens if is_token_valid(token)]

    if not valid_numbers:
        return "", "", False

    last_two_numbers = valid_numbers[-2:]

    if len(last_two_numbers) == 1:
        id_number = last_two_numbers[0]
        return id_number, "", True

    id_number = last_two_numbers[0]
    current_revision_number = last_two_numbers[1]

    last_two_digits = id_number[-2:] if len(id_number) >= 2 else id_number

    formatted_string = "refs/changes/{}/{}/{}".format(last_two_digits, id_number, current_revision_number)

    return formatted_string, current_revision_number, False


def get_absolute_path(path):
    if os.path.isabs(path):
        # 如果是绝对路径，直接返回完整的绝对路径
        return os.path.abspath(path)
    else:
        # 如果是相对路径，将当前路径和相对路径拼接成绝对路径
        current_dir = os.getcwd()
        absolute_path = os.path.join(current_dir, path)
        return os.path.abspath(absolute_path)


def parse_cmd_args():
    parser = argparse.ArgumentParser(description="ogki gki oki build tools")
    parser.add_argument('-p', '--path', type=str, help='default dirname', default=get_absolute_path('google'))
    parser.add_argument('-u', '--url', type=str, help='manifest url',
                        default='https://partner-android.googlesource.com/kernel/manifest')
    parser.add_argument('-s', '--common', type=str, help='aosp source code', default='')
    parser.add_argument('-b', '--branch', type=str, help='manifest branch', default='common-android15-6.6-oppo')
    parser.add_argument('-m', '--manifest', type=str, help='manifest file name', default='')
    parser.add_argument('-e', '--ext', type=str, help='repo special ext command', default='')
    parser.add_argument('-t', '--type', type=str, help='branch private or public', default='partner')
    parser.add_argument('-o', '--out', type=str, help='dist output', default='oplus')
    parser.add_argument('-z', '--zip', type=str, help='zip file name', default='dist.tar.gz')
    parser.add_argument('-n', '--bug_number', type=str, help='Google issue number', default='000000000')
    parser.add_argument('-g', '--gerrit_number', type=str, help='Google gerrit number', default='')
    parser.add_argument('-c', '--config_info', type=str, help='config info', default=get_tools_path('config_info.xlsx'))
    parser.add_argument('-r', '--review_url', type=str, help='review gerrit url', default='')
    parser.add_argument('-w', '--gerrit_url', type=str, help='web gerrit url', default='')
    parser.add_argument('-a', '--auto', action='store_true', help='auto get info', default=False)
    parser.add_argument('-l', '--log', type=str, help='out put log dir', default='')
    parser.add_argument('-x', '--max', type=str, help='max change list', default='10')
    parser.add_argument('--refs', type=str, help='refs change', default='')
    parser.add_argument('--out_abi', type=str, help='out abi', default='out_abi/kernel_aarch64/dist')
    parser.add_argument('--user', type=str, help='default user', default='')
    parser.add_argument('--new', type=str, help='new build', default='true')
    parser.add_argument('--artifactory', type=str, help='default user', default='')
    parser.add_argument('--action', type=str, help='default download', default='build')
    parser.add_argument('--oem_ki', type=str, help='build ogki or gki', default='')
    parser.add_argument('--commit_lists_id', type=str, help='commit_lists_id', default='commit_lists_id.xlsx')
    parser.add_argument('--commit_lists_web', type=str, help='commit_lists_id',
                        default=get_tools_path('commit_lists_web.txt'))

    args = parser.parse_args()

    args.path = get_absolute_path(args.path)
    args.type, oem_ki = check_if_partner(args.url)
    if not args.oem_ki:
        args.oem_ki = oem_ki
    branch = extract_android_version(args.branch)
    current_path = process_path(args.path, args.type, branch)  # 输出：google-aosp-android15-6.6
    print("current_path:", current_path)
    create_symlink(args.path, current_path)
    args.path = current_path

    args.out = os.path.join(args.path, branch)

    args.zip = os.path.join(args.out, branch + ".tar.gz")
    args.common = os.path.join(args.path, "common")
    args.log = os.path.join(args.out, "log")
    args.out_abi = os.path.join(args.path, "out_abi/kernel_aarch64*/dist")
    args.commit_lists_id = os.path.join(args.out, "commit_lists_id.xlsx")

    if args.type.lower() == "partner" and args.action.lower() == "build":
        args.review_url = "https://partner-android-review.googlesource.com"
        args.gerrit_url = "https://partner-android.googlesource.com/kernel-oem/oppo/common"
    else:
        args.review_url = "https://android-review.googlesource.com"
        args.gerrit_url = "https://android.googlesource.com/kernel/common"

    args.refs, revision_number, args.auto = parse_gerrit_number(args.gerrit_number)
    # print("input {}\n parse：{}\n rebase state：{}\n".format(args.gerrit_number, args.refs, args.auto))

    if not args.gerrit_number:
        print("now your gerrit number is null,you must input a gerrit number like 3249938 or 3249938/1")
        # sys.exit(1)

    for key, value in vars(args).items():
        print(f"{key}: {value}")

    if args.gerrit_number and not revision_number:
        message = (
            "cannot find revision number by command. \n"
            "You must input the full form like '3249938/1'. The form '3249938' is not supported."
        )
        print(message)
        sys.exit(1)
    # disable warning
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    return args


def delete_matching_directories(base_directory, patterns):
    """
    在基目录中查找匹配指定正则表达式模式的所有目录，
    如果找到匹配的目录，则删除它们。

    :param base_directory: 要检查的基目录路径
    :param patterns: 模糊匹配的正则表达式模式列表
    """
    if not os.path.isdir(base_directory):
        print(f"The directory {base_directory} does not exist or is not a directory.")
        return

    # 列出基目录中的所有条目
    entries = os.listdir(base_directory)

    # 记录所有匹配的目录
    matched_directories = []

    for entry in entries:
        entry_path = os.path.join(base_directory, entry)
        # 检查是否是目录或符号链接目录
        if os.path.isdir(entry_path) or (os.path.islink(entry_path)):
            for pattern in patterns:
                if re.search(pattern, entry):
                    matched_directories.append(entry_path)
                    print(f"Matched pattern: {pattern.pattern} in {entry}")
                    # 不需要 break，这样会继续检查其他模式

    # 删除所有匹配的目录
    for dir_path in matched_directories:
        try:
            if os.path.islink(dir_path):
                delete_symlink(dir_path)
            else:
                shutil.rmtree(dir_path)
            print(f"Deleted directory: {dir_path}")
        except Exception as e:
            print(f"Failed to delete directory {dir_path}: {e}")

    if not matched_directories:
        print("No matching directories found.")


def init_work_env(args, path, dist):
    if not os.path.exists(path):
        os.makedirs(path)

    out = os.path.join(args.path, "out")
    abi = os.path.join(args.path, "out_abi")

    if args.new.lower() != "false" and args.action.lower() == "build":
        print("this is a clean build,default delete out", args.new)
        if os.path.exists(out):
            print('\ndelete ', out)
            shutil.rmtree(out)

        if os.path.exists(abi):
            print('\ndelete ', abi)
            shutil.rmtree(abi)

    if os.path.exists(dist):
        print('\ndelete ', dist)
        shutil.rmtree(dist)

    patterns = [
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+\b"),
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+_abogki\d+\b"),
        re.compile(r"android\d+-\d+\.\d+_r\d+_abogki\d+\b"),
        re.compile(r"android\d+-\d+\.\d+_r\d+\b")
    ]

    delete_matching_directories(args.path, patterns)

    if not os.path.exists(dist):
        os.makedirs(dist)

    if not os.path.exists(args.log):
        os.makedirs(args.log)

    os.chdir(path)


def run_repo_init_command(args, path, url, branch, manifest, ext):
    command = [repo_cmd, 'init']

    if url:
        command.extend(['-u', url])
    if branch:
        command.extend(['-b', branch])
    if manifest:
        command.extend(['-m', manifest])
    if ext:
        command.extend([ext])

    print('\ncommand', command)
    subprocess.run(command)


def run_repo_sync_command(args):
    command = "{} sync -c -j4 --prune --no-repo-verify --force-sync".format(repo_cmd)
    print('\ncommand', command)
    subprocess.run(command, shell=True)


def build_kernel_cmd(args, number):
    build_number = number.split("/")[0]
    command = []
    if build_number:
        if args.oem_ki.lower() == "ogki":
            os.environ['BUILD_NUMBER'] = "ogki" + build_number
        elif args.oem_ki.lower() == "gki":
            os.environ['BUILD_NUMBER'] = build_number
        print('BUILD_NUMBER=ab{}{}'.format(args.oem_ki.lower(), build_number))

    command.extend(['tools/bazel', 'run', '--config=release', '//common:kernel_aarch64_abi_dist'])
    print('number {} build_number {}'.format(number, build_number))
    env = os.environ.copy()  # 复制当前环境变量
    command_string = subprocess.list2cmdline(command)
    print(command_string)
    subprocess.run(command, env=env)  # 将复制的环境变量传递给subprocess.run()


def tar_dir_to_gz(zip_file, out_abi, log_file):
    if os.path.exists(zip_file):
        os.remove(zip_file)

    out_abis = find_file_with_wildcard(out_abi)

    if out_abis and len(out_abis) == 1:
        out_abi = out_abis[0]
        print('\nout_abi:', out_abi)
        # Get the parent directory and the base name of out_abi
        parent_dir, base_name = os.path.split(out_abi)

        # Create the tar command with the correct options
        command = ['tar', '-czvf', zip_file, '-C', parent_dir, base_name]
        print('\ncommand', command)
        with open(log_file, 'w') as f:
            result = subprocess.run(command, shell=False, stdout=f, stderr=subprocess.STDOUT)
            # print(result.stdout)
        if os.path.exists(zip_file):
            print(f"Archive {zip_file} created successfully.")
        else:
            print(f"Failed to create archive {zip_file}. Check log file {log_file} for details.")
    else:
        print(f"Ignored files: {out_abis}")


def tar_dir_to_gzip9(zip_file, out_abi, log_file):
    if os.path.exists(zip_file):
        os.remove(zip_file)

    out_abis = find_file_with_wildcard(out_abi)

    if out_abis and len(out_abis) == 1:
        out_abi = out_abis[0]
        print('\nout_abi:', out_abi)

        # Get the parent directory and the base name of out_abi
        parent_dir, base_name = os.path.split(out_abi)

        # Prepare commands
        tar_command = ['tar', '-c', '-C', parent_dir, base_name]
        gzip_command = ['gzip', '-9']

        print('\ntar command:', tar_command)
        print('gzip command:', gzip_command)

        with open(log_file, 'w') as log_f, open(zip_file, 'wb') as gz_file:
            # Run the tar command and pipe its output to gzip
            tar_process = subprocess.Popen(tar_command, stdout=subprocess.PIPE, stderr=log_f)
            gzip_process = subprocess.Popen(gzip_command, stdin=tar_process.stdout, stdout=gz_file, stderr=log_f)

            # Ensure tar_process's stdout is closed in gzip_process
            tar_process.stdout.close()

            # Wait for the processes to complete
            tar_return_code = tar_process.wait()
            gzip_return_code = gzip_process.wait()

        # Check if the gzip process was successful
        if gzip_return_code == 0:
            print(f"Archive {zip_file} created successfully.")
        else:
            print(f"Failed to create archive {zip_file}. Check log file {log_file} for details.")
    else:
        print(f"Ignored  files: {out_abis}")


def find_file_with_wildcard(pattern):
    matches = glob.glob(pattern)
    if matches:
        return matches
    return None


def copy_files_to_dist(src, local_file_paths, dst):
    if not os.path.exists(dst):
        os.makedirs(dst)
    for file_name, new_name in local_file_paths:
        # src_path = os.path.join(src, file_name)
        src_paths = find_file_with_wildcard(os.path.join(src, file_name))
        dst_path = os.path.join(dst, new_name if new_name else file_name)
        if src_paths and len(src_paths) == 1:

            # print(f"Matched {len(src_paths)} files:")
            for src_path in src_paths:
                # print(src_path)
                print('copy {} to {}'.format(src_path, dst_path))
                shutil.copy(src_path, dst_path)
        else:
            print('{} ignore'.format(src_paths))


def get_build_info(output_path, repo_path='.'):
    try:
        result = subprocess.run(['repo', 'manifest', '-r', '-o', output_path], cwd=repo_path, stdout=subprocess.PIPE,
                                check=True)
        return result.stdout.strip().decode('utf-8')
    except subprocess.CalledProcessError as e:
        return e.stderr.strip().decode('utf-8')


def download_gerrit_info_request(args, file_name, query, id_type="changes", branch='android16-6.12'):
    project = 'kernel/common'
    if not args.review_url:
        raise ValueError("review_url is required")
    if not query:
        raise ValueError("query is required")

    if id_type == "detail":
        url = "{}/changes/{}/detail".format(args.review_url, query)
    elif id_type == "revision":
        url = "{}/changes/?q=project:{}+{}&o=ALL_REVISIONS".format(args.review_url, project, query)
    elif id_type == "changeid":
        url = "{}/changes/?q=change:{}+branch:{}".format(args.review_url, query, branch)
    else:
        url = "{}/changes/?q=project:{}+{}".format(args.review_url, project, query)
    print('url:{}\n'.format(url))

    payload = {}
    files = {}
    headers = {}

    response = requests.get(url, headers=headers, data=payload, files=files, verify=False)
    # print(response.text)
    with open(file_name, 'w') as file:
        file.write(response.text)
    # print("\nResponse text already written to file:", file_name)


def pretty_print_json(input_file, output_file):
    try:
        with open(input_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
    except json.JSONDecodeError as e:
        print(f"JSON 解析错误: {e}")
        raise

    with open(output_file, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def download_gerrit_info_analysis(args, query, id_type):
    download_json = os.path.join(args.log, "{}_{}_download.json".format(query, id_type))
    valid_json = os.path.join(args.log, "{}_{}_valid.json".format(query, id_type))
    output_json = os.path.join(args.log, "{}_{}_output.json".format(query, id_type))
    download_gerrit_info_request(args, download_json, query, id_type)
    filter_gerrit_info(download_json, valid_json)
    pretty_print_json(valid_json, output_json)
    return output_json


def session_download_gerrit_info(session, args, file_name, query):
    url = "{}/changes/?q={}".format(args.review_url, query)
    print('\nurl:', url)

    response = session.get(url, verify=False)
    # print(response.text)
    with open(file_name, 'w') as file:
        file.write(response.text)

    print("\nResponse text already written to file:", file_name)
    """"""


def find_default_user_info(args, email, remote_name, file_path='config_info.xlsx'):
    # 加载Excel工作簿
    print('\nfile_path', file_path)
    author_password = ""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 获取各列的索引
    column_names = [cell.value.strip() for cell in sheet[1]]
    # print(column_names)
    email_index = column_names.index("Committer_Email") + 1
    # print(email_index)
    password_index = column_names.index("password") + 1

    for row in range(2, sheet.max_row + 1):
        author_email = sheet.cell(row=row, column=email_index).value
        author_password = sheet.cell(row=row, column=password_index).value
        # print(author_email)
        # print(email)
        if author_email == email:
            break
    else:
        print("{} not found in the Excel file.".format(email))

    print(author_password)
    return author_password


def login_download_gerrit_info(args, file_name, query):
    # 你的Gerrit用户名和密码
    username = 'xxx'
    password = 'xxx'

    login_url = "{}/login".format(args.review_url)
    print('\nlogin_url:', login_url)
    password = find_default_user_info(args, args.user, args.type, args.config_info)
    # 创建一个会话对象
    session = requests.Session()
    # 进行登录，获取cookie
    login_payload = {
        # 'username': username,
        'username': args.user,
        'password': password
    }

    login_response = session.post(login_url, data=login_payload, verify=False)

    # 检查是否登录成功
    if login_response.status_code == 200:
        print("Login successful")
        session_download_gerrit_info(session, args, file_name, query)
    else:
        print("Login failed")


def download_gerrit_info(args, file_name, query):
    if args.type.lower() == "partner":
        print("to do later current use default")
        # login_download_gerrit_info(args, file_name, query)
    else:
        download_gerrit_info_request(args, file_name, query)


def filter_gerrit_info(input_file, output_file):
    with open(input_file, 'r') as f:
        lines = f.readlines()

    valid_jsons = []
    for line in lines:
        try:
            json.loads(line)
            valid_jsons.append(line)
        except json.JSONDecodeError:
            # print('\n ignore invalid line \n\n{}\nin {}\n'.format(line, input_file))
            continue

    with open(output_file, 'w') as f:
        f.writelines(valid_jsons)


def parse_gerrit_info_from_json_to_excel(args, file_name, out_put, overwrite=True):
    extracted_data = []

    with open(file_name, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
            for item in data:

                branch = item['branch']
                virtual_id_number = str(item['virtual_id_number'])
                current_revision_number = str(item['current_revision_number'])

                last_two_digits = str(virtual_id_number)[-2:]
                refs = "refs/changes/{}/{}/{}".format(last_two_digits, virtual_id_number, current_revision_number)

                fetch_url = "{} {}".format(args.gerrit_url, refs)

                cherry_pick_command = "git fetch {} && git cherry-pick FETCH_HEAD ".format(fetch_url)
                checkout_command = "git fetch {} && git checkout FETCH_HEAD ".format(fetch_url)

                if args.type.lower() == "private":
                    push = "git push partner HEAD:refs/for/{}".format(branch)
                else:
                    push = "git push aosp HEAD:refs/for/{}".format(branch)

                extracted_item = {
                    'project': item['project'],
                    'branch': item['branch'],
                    'virtual_id_number': virtual_id_number,
                    'current_revision_number': current_revision_number,
                    'refs': refs,
                    'status': item['status'],
                    'cherry_pick_command': cherry_pick_command,
                    'checkout_command': checkout_command,
                    'push': push,
                    'change_id': item['change_id'],
                    'subject': item['subject']
                }
                extracted_data.append(extracted_item)
        except json.JSONDecodeError:
            print("by pass for error line")

    # 转换为DataFrame
    df = pd.DataFrame(extracted_data)

    if not overwrite:
        try:
            existing_df = pd.read_excel(out_put)
            merged_df = pd.concat([existing_df, df]).drop_duplicates()
            df = merged_df
        except FileNotFoundError:
            pass

    # 输出到Excel文件
    with pd.ExcelWriter(out_put, engine='openpyxl', mode='a' if not overwrite else 'w') as writer:
        df.to_excel(writer, index=False)
    print("data already write to {}".format(out_put))
    return df, df['status'].iloc[-1] if len(df) > 0 else None


def find_checkout_command_and_run(args, file_path, gerrit_id):
    """
    Reads an Excel file and finds the checkout command for a given gerrit_id.

    :param args: Arguments containing the common directory.
    :param file_path: Path to the Excel file.
    :param gerrit_id: gerrit_id to search for.
    :return: The last branch value corresponding to the matched gerrit_id.
    """
    branches = []
    try:
        # Read the Excel file
        df = pd.read_excel(file_path)

        # Ensure virtual_id_number is of type str
        df['virtual_id_number'] = df['virtual_id_number'].astype(str)

        # Ensure gerrit_id is also of type str
        gerrit_id = str(gerrit_id)

        # Filter the rows where virtual_id_number matches the given gerrit_id
        matched_rows = df[df['virtual_id_number'] == gerrit_id]

        # Check if any rows matched
        if not matched_rows.empty and len(matched_rows.columns) > 1:
            # Print the checkout_command column's content of matched rows
            for index, row in matched_rows.iterrows():
                command = row['checkout_command']
                branch = row['branch']
                # Print the checkout command and the corresponding branch
                print("Checkout Command: {}".format(command))
                print("Branch: {}".format(branch))
                # Execute the command
                current_path = os.getcwd()
                os.chdir(args.common)
                run_git_command(command)
                os.chdir(current_path)
                # Append the branch to the list
                branches.append(branch)
        else:
            print("No matching virtual_id_number found for gerrit_id: {}".format(gerrit_id))

    except Exception as e:
        print("An error occurred: {}".format(e))

    # Print all matched branches
    for b in branches:
        print("Matched branch: {}".format(b))

    # Return the last branch if there are any branches matched
    return branches[-1] if branches else None


def tee_output_to_file_and_terminal(command, log_file):
    with open(log_file, 'w') as log:
        # 使用 Popen 而不是 run，因为 Popen 给我们提供了更多的控制
        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,  # 行缓冲
            universal_newlines=True  # 自动解码
        )

        # 逐行读取 stdout 和 stderr
        while True:
            output = process.stdout.readline()
            if output == '' and process.poll() is not None:
                break
            if output:
                sys.stdout.write(output)
                log.write(output)

        while True:
            error = process.stderr.readline()
            if error == '' and process.poll() is not None:
                break
            if error:
                sys.stderr.write(error)
                log.write(error)

        # 确保所有子进程流被关闭
        process.stdout.close()
        process.stderr.close()

        # 等待子进程的退出并得到返回码
        return_code = process.wait()
        return return_code


def handle_artifactory_path(artifactory_path, config_file_path):
    if not artifactory_path:
        print("artifactory_path is null")
        return True

    # Get the basename to check if it's a directory or a file
    basename = os.path.basename(artifactory_path)

    # Determine if it's a file based on the basename
    if '.' in basename:
        # Split the basename to check the extension part
        name, ext = os.path.splitext(basename)
        ext_len = len(ext) - 1
        # Typically, extensions are 1 to 4 characters long
        if 1 <= ext_len <= 4:
            is_file = True
        else:
            is_file = False
    else:
        is_file = False

    # Determine if it's a file based on the basename
    if is_file:
        # Get the directory path
        directory = os.path.dirname(artifactory_path)

        # Create the directory if it doesn't exist
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            print("Directory created: {}.".format(directory))

        # Create an empty file
        with open(artifactory_path, 'w'):
            pass
        print("File created: {}.".format(artifactory_path))

        # Copy the configuration file to the specified path
        shutil.copy(config_file_path, artifactory_path)
        print("File copied from {} to {}.".format(config_file_path, artifactory_path))
    else:
        # It's a directory if there is no file extension
        if not os.path.exists(artifactory_path):
            os.makedirs(artifactory_path)
            print("Directory created: {}.".format(artifactory_path))

        # Copy the configuration file to the directory
        destination_path = os.path.join(artifactory_path, os.path.basename(config_file_path))
        shutil.copy(config_file_path, destination_path)
        print("File copied from {} to directory {}.".format(config_file_path, artifactory_path))


def upload_oem_ki_build(args, out, log_file):
    command = [
        "python3", get_tools_path("ogki_gki_artifactory.py"),
        "-t", "upload",
        "-k", args.oem_ki.upper(),
        "-i", args.out,
        "-o", args.out
        # "-f"
    ]

    #print('\ncommand:', command)
    command_string = subprocess.list2cmdline(command)
    print(command_string)

    try:
        return_code = tee_output_to_file_and_terminal(command, log_file)

        if return_code == 0:
            print('Successfully run command {}'.format(command))
        else:
            print('Failed to run command {}'.format(command))

    except Exception as e:
        print("run command error: {}".format(e))

    # Example:
    # Replace args.artifactory with the path you want to pass, e.g., 'test/file.txt' or 'test/directory'
    # Replace args.config_file with the path to the configuration file, e.g., 'out/config.txt'
    config = os.path.join(out, 'config.txt')
    handle_artifactory_path(args.artifactory, config)


def git_format_patch(args, common, out, commit_hash):
    """
    Generate patches from the specified commit hash to the latest commit.

    :param common: The path to the git repository.
    :param out: The directory where the patch files will be generated.
    :param commit_hash: The commit hash to start generating patches from.
    :param args: Arguments containing the common directory.
    """

    if commit_hash:
        command = "git -C {} format-patch {}..HEAD -o {}".format(common, commit_hash, out)
    else:
        command = "git -C {} format-patch -{} -o {}".format(common, args.max, out)
    print('\nCommand:', command)
    try:
        subprocess.run(command, shell=True, check=True)
        print("Successfully generated patches.")
    except subprocess.CalledProcessError as e:
        print("Failed to execute command: {}".format(e))


def get_max_commits_list(repo_path, num_commits=3):
    # Git log command to get commits in a specified format
    cmd = "git -C {} log --pretty=format:'%H|%ae|%an|%ce|%cn|%s' --no-merges -{}".format(repo_path, num_commits)

    print(f"Command: {cmd}")
    # 将命令列表转换为字符串
    # command_string = subprocess.list2cmdline(cmd)
    # print(command_string)

    # Execute the command and get the output
    result = subprocess.check_output(cmd, shell=True).decode('utf-8')
    # Split the output into lines
    lines = result.strip().splitlines()
    # Split each line into fields
    data = [re.split(r'\|', line) for line in lines]

    # Predefine all expected columns
    columns = ['Commit_Hash', 'Author_Email', 'Author_Name', 'Committer_Email', 'Committer_Name', 'Subject',
               'project', 'branch', 'virtual_id_number', 'current_revision_number', 'refs',
               'status', 'cherry_pick_command', 'checkout_command', 'push', 'change_id', 'subject2']

    # Create an empty DataFrame with all possible columns
    df = pd.DataFrame(columns=columns)

    # Add rows to the DataFrame, filling missing columns with None or NaN
    rows = []
    for row in data:
        row_dict = {col: (row[index] if index < len(row) else None) for index, col in enumerate(columns[:len(row)])}
        rows.append(row_dict)

    # Use pd.concat to add rows
    df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True, sort=False)

    return df


def write_commits_list_to_excel(df, output_file):
    df.to_excel(output_file, index=False)


def get_valid_commit_lists(args, file_path):
    merged_hash = ""
    # Load the Excel file
    df = pd.read_excel(file_path)

    # Find the column index for "project"
    project_col_index = df.columns.get_loc("project") + 1  # Excel column indices are 1-based

    # Find the index of the "Commit_Hash" column
    commit_hash_col_index = df.columns.get_loc("Commit_Hash")

    # Get the values of the "Commit_Hash" column
    commit_hash_column = df.iloc[:, commit_hash_col_index]

    # List to collect unprocessed commit hashes
    unprocessed_commit_hashes = []

    # Iterate over the items in the "Commit_Hash" column and process each commit_hash
    for index, commit_hash in enumerate(commit_hash_column):
        print(commit_hash)
        current_change_excel = os.path.join(args.log, "download_commit_file.xlsx")

        current_change_filtered_path = download_gerrit_info_analysis(args, commit_hash, 'changes')

        df, status = parse_gerrit_info_from_json_to_excel(args, current_change_filtered_path, current_change_excel)

        # If status does not equal "MERGED", write the df values to the corresponding row
        # starting from the project column
        if status is not None and status.upper() != "MERGED":
            print("Status is append, exiting...")
            # Determine the row index (1-based)
            row_index = index + 2  # Since we read from the first row, we need to add 2

            # Append unprocessed commit hashes
            unprocessed_commit_hashes.append(commit_hash)

            # Write the df values to the same row starting from the project column
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay',
                                engine='openpyxl') as writer:  # Use append mode
                for col_index, value in enumerate(df.values[0], start=project_col_index):
                    writer.sheets['Sheet1'].cell(row=row_index, column=col_index, value=value)
        else:
            print("Status is merged, exiting...")
            merged_hash = commit_hash
            break

    # If the loop exits without setting merged_hash, it means no "MERGED" status was found.
    # Print and return the unprocessed commit hashes
    if not merged_hash:
        unprocessed_commit_hashes_str = ", ".join(unprocessed_commit_hashes)
        print("Unprocessed commit hashes:", unprocessed_commit_hashes_str)
        return unprocessed_commit_hashes_str
    else:
        print("Merged hash found:", merged_hash)
        return merged_hash


def set_git_config(name, email):
    # 设置全局用户名
    subprocess.run(["git", "config", "--global", "user.name", name])
    # 设置全局用户邮箱
    subprocess.run(["git", "config", "--global", "user.email", email])


def switch_to_branch(branch_name, remote_name='aosp'):
    # 获取当前分支名
    current_branch = subprocess.getoutput('git rev-parse --abbrev-ref HEAD')

    # 检查指定分支是否存在
    branches = subprocess.getoutput('git branch -a').split('\n')

    # 完全匹配的分支名
    matched_branches = [b.strip() for b in branches if b.strip().endswith('{}/{}'.format(remote_name, branch_name))]

    # 检查本地是否已经存在该分支
    local_branch_exists = branch_name in subprocess.getoutput('git branch')

    if matched_branches:
        if local_branch_exists:
            if current_branch != branch_name:
                subprocess.run(['git', 'checkout', branch_name])
        else:
            subprocess.run(['git', 'checkout', '-b', branch_name, 'remotes/{}/{}'.format(remote_name, branch_name)])
    else:
        print('Branch {} does not exist under remote {}.'.format(branch_name, remote_name))


def run_git_command(command):
    print('\ncommand', command)
    # 执行命令
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 获取命令输出和错误信息
    output, error = process.communicate()
    # 返回命令的输出和错误信息
    if process.returncode == 0:
        print('Successfully run command  {}'.format(command))
        message = output.decode('utf-8')
        if message:
            print("Output:\n", message)
    else:
        print('Failed to  run command  {}'.format(command))
        print("decode:\n", error.decode('utf-8'))

    return


def git_push(branch_name, remote_name='aosp'):
    command = ['git', 'push', remote_name, 'HEAD:refs/for/{}'.format(branch_name)]

    # 打印执行的命令
    print("Executing command: {}".format(command))

    result = subprocess.run(command)
    if result.returncode == 0:
        print('Successfully pushed to {} on branch {}'.format(remote_name, branch_name))
    else:
        print('Failed to push to {} on branch {}'.format(remote_name, branch_name))


def clear_gitcookies_script(shell_script):
    """
    Create a shell script to clear the contents of ~/.gitcookies.
    :param shell_script: The path to the shell script to generate.
    """
    with open(shell_script, 'w') as script:
        # Write shell script header
        script.write("#!/bin/bash\n\n")
        script.write("# This script clears the contents of ~/.gitcookies and executes a command\n\n")

        # Write command to clear .gitcookies
        script.write("echo 'Clearing the contents of ~/.gitcookies...'\n")
        script.write("> ~/.gitcookies\n")

    print("{} \ncreated to clear ~/.gitcookies and execute a command.".format(shell_script))


def append_command_to_script(shell_script, command):
    """
    Append a command to the shell script.
    :param shell_script: The path to the shell script.
    :param command: The command to append to the shell script.
    """
    with open(shell_script, 'a') as script:
        # Append the command
        script.write("\n# Execute the provided command\n")
        script.write("echo 'Executing command: {}'\n".format(command))
        script.write("{}\n".format(command))


def execute_shell_script(shell_script, output_file):
    """
    Execute the generated shell script.
    :param shell_script: The path to the shell script to execute.
    :param output_file: The file to write both stdout and stderr logs.
    """
    try:
        # Make the shell script executable
        subprocess.run("chmod +x {}".format(shell_script), shell=True, check=True)

        # Open the output_file in append mode
        with open(output_file, 'a') as file:
            # Execute the shell script and capture the output
            result = subprocess.run("{}".format(shell_script), shell=True, check=True, stdout=subprocess.PIPE)
            output = result.stdout.decode('utf-8')

            # Write the stdout and stderr to the file
            file.write("Shell script executed successfully:\n")
            file.write(output)

    except subprocess.CalledProcessError as e:
        # If an error occurs, write the stderr to the file
        with open(output_file, 'a') as file:
            file.write("Error occurred while executing shell script:\n")
            file.write(e.stderr)


def run_shell_command_with_script(args, shell_script, command):
    """
    Perform the task of clearing .gitcookies and executing a command using a shell script.
    :param shell_script: The path to the shell script to generate.
    :param command: The command to execute.

   :param args: Arguments containing the common directory.
    """
    clear_gitcookies_script(shell_script)
    append_command_to_script(shell_script, command)
    execute_shell_script(shell_script, os.path.join(args.log, "custom_script.log"))


def default_set_gerrit_config_info(args, email, remote_name, file_path='config_info.xlsx'):
    # 加载Excel工作簿
    print('\nfile_path', file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 获取各列的索引
    column_names = [cell.value.strip() for cell in sheet[1]]
    # print(column_names)
    email_index = column_names.index("Committer_Email") + 1
    # print(email_index)
    aosp_index = column_names.index("aosp_Authenticate") + 1
    # print(aosp_index)
    partner_index = column_names.index("partner_Authenticate") + 1

    for row in range(2, sheet.max_row + 1):
        author_email = sheet.cell(row=row, column=email_index).value
        authenticate_aosp = sheet.cell(row=row, column=aosp_index).value
        authenticate_partner = sheet.cell(row=row, column=partner_index).value
        # print(author_email)
        # print(email)
        if author_email == email:
            if remote_name == 'partner':
                if authenticate_partner:
                    # print("Partner Authenticate for {}: {}".format(email, authenticate_partner))
                    run_shell_command_with_script(args, os.path.join(args.log, "custom_script.sh"),
                                                  authenticate_partner)

                else:
                    print("No Partner Authenticate found for {}".format(email))
            else:
                if authenticate_aosp:
                    run_shell_command_with_script(args, os.path.join(args.log, "custom_script.sh"), authenticate_aosp)
                else:
                    print("No AOSP Authenticate found for {}".format(email))
            break
    else:
        print("{} not found in the Excel file.".format(email))


def rebase_valid_commits_list_and_push(args, file_path):
    # Read Excel file
    df = pd.read_excel(file_path)

    # Reverse order of rows
    df = df.iloc[::-1]

    # Iterate over rows
    for index, row in df.iterrows():
        if pd.notna(row['status']) and row['status'] != '':
            # Print each column value in the row
            # for col in df.columns:
            #    print("row {} {}: {}".format(index, col, row[col]))
            # print("\n")  # Newline for better readability
            committer_email = row['Committer_Email']
            committer_name = row['Committer_Name']
            branch = row['branch']
            set_git_config(committer_name, committer_email)
            default_set_gerrit_config_info(args, committer_email, args.type, args.config_info)
            switch_to_branch(branch, args.type)
            run_git_command(row['cherry_pick_command'])
            git_push(branch, args.type)
        # else:
        #    print("Empty status in the row {}".format(index))


def git_reset_and_pull(original_hash):
    """
    Function to reset git to a specific commit hash, pull the latest changes,
    and return the latest commit hash.
    :param original_hash: The hash to reset the git repository to.
    :return: The latest commit hash after pulling the latest changes.
    """
    try:
        # Step 2: Force reset to the specified commit hash
        result = subprocess.run(["git", "reset", "--hard", original_hash], check=True, stdout=subprocess.PIPE)
        print("Successfully reset to {}".format(original_hash))

        # Step 3: Pull the latest changes
        result = subprocess.run(["git", "pull"], check=True, stdout=subprocess.PIPE)
        print("Successfully pulled the latest changes")

        # Step 4: Get the latest commit hash
        result = subprocess.run(["git", "rev-parse", "HEAD"], check=True, stdout=subprocess.PIPE)
        latest_hash = result.stdout.strip().decode('utf-8')
        print("The latest commit hash is {}".format(latest_hash))

        return latest_hash

    except subprocess.CalledProcessError as e:
        print("An error occurred:\n{}".format(e.stderr))
        raise


def update_valid_commits_list_and_push(args, excel, branch, merged_hash):
    current_path = os.getcwd()
    os.chdir(args.common)

    print("switch branch: {}".format(branch))
    switch_to_branch(branch, args.type)
    # Example usage
    latest_hash = git_reset_and_pull(merged_hash)
    print("Updated to the latest hash: {}".format(latest_hash))

    rebase_valid_commits_list_and_push(args, excel)
    os.chdir(current_path)

    return latest_hash


def get_current_revision(args):
    current_path = os.getcwd()
    os.chdir(args.common)
    fetch_url = "{} {}".format(args.gerrit_url, args.refs)
    checkout_command = "git fetch {} && git checkout FETCH_HEAD ".format(fetch_url)
    # print('\n command', checkout_command)
    run_git_command(checkout_command)
    os.chdir(current_path)
    return ""


def get_latest_revision(args):
    valid_download_excel = os.path.join(args.log, "valid_download.xlsx")

    output_json = download_gerrit_info_analysis(args, args.gerrit_number, 'changes')
    parse_gerrit_info_from_json_to_excel(args, output_json, valid_download_excel)
    branch = find_checkout_command_and_run(args, valid_download_excel, args.gerrit_number)
    print('\nbranch', branch)
    return branch


def pull_to_revision(args, max_commits_list_excel):
    merged_hash = ""

    if args.auto:
        branch = get_latest_revision(args)
    else:
        branch = get_current_revision(args)

    if args.type.lower() == "aosp":
        df = get_max_commits_list(args.common, args.max)
        write_commits_list_to_excel(df, max_commits_list_excel)
        merged_hash = get_valid_commit_lists(args, max_commits_list_excel)

    print("merged_hash {}  branch: {}".format(merged_hash, branch))
    return merged_hash, branch


def run_repo_command(args):
    try:
        default_set_gerrit_config_info(args, args.user, args.type, args.config_info)
        run_repo_init_command(args, args.path, args.url, args.branch, args.manifest, args.ext)
        run_repo_sync_command(args)

    except Exception as e:
        # 打印异常信息（可选）
        print("repo error is: {}".format(e))
        sys.exit(1)
    return


def extract_and_save_numbers(input_file, output_file, headers):
    # 读取文件内容
    with open(input_file, 'r') as file:
        content = file.read()

    # 使用正则表达式提取所有长度一致的数字字符
    numbers = re.findall(r'\b\d{7}\b', content)

    # 去重并排序
    unique_numbers = sorted(set(numbers))

    # 根据文件后缀决定写入 CSV 还是 Excel
    if output_file.lower().endswith('.csv'):
        with open(output_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)  # 写入表头
            for number in unique_numbers:
                writer.writerow([number] + [''] * (len(headers) - 1))
    elif output_file.lower().endswith('.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Commit_IDs"
        ws.append(headers)  # 写入表头
        for number in unique_numbers:
            ws.append([number] + [''] * (len(headers) - 1))
        wb.save(output_file)
    else:
        raise ValueError("Output file must have a .csv or .xlsx extension")


def get_latest_gerrit_commit_info(repo_path, excel_file_path):
    # Change directory to the repository path
    original_cwd = subprocess.getoutput('pwd')
    subprocess.run(['cd', repo_path], shell=True)

    # Get the latest commit details
    commit_hash = subprocess.getoutput('git rev-parse HEAD').strip()
    committer_name = subprocess.getoutput('git show -s --format=%cn').strip()
    committer_email = subprocess.getoutput('git show -s --format=%ce').strip()
    subject = subprocess.getoutput('git show -s --format=%s').strip()
    change_id = subprocess.getoutput('git show -s --format=%b | grep Change-Id | awk \'{print $2}\'').strip()

    # Change back to the original directory
    subprocess.run(['cd', original_cwd], shell=True)

    # Create a dictionary with the commit details
    commit_info = {
        'Commit_Hash': [commit_hash],
        'Committer_Name': [committer_name],
        'Committer_Email': [committer_email],
        'Subject': [subject],
        'Change_Id': [change_id]
    }

    # Convert the dictionary to a DataFrame
    df = pd.DataFrame(commit_info)

    # Save the DataFrame to an Excel file
    # excel_file_path = 'latest_commit_info.xlsx'
    df.to_excel(excel_file_path, index=False)

    return excel_file_path


def read_excel_changes(file_path):
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 提取 change 列的值并转换为字符串
    changes = df.get("src_change", []).astype(str)

    return df, changes


def read_excel_change_ids(file_path):
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 提取 change 列的值并转换为字符串
    change_ids = df.get("change_id", []).astype(str)

    return df, change_ids


def get_commit_lists_from_file(args, commit_lists_web, output_file_excel):
    headers = ['src_change', 'src_url', 'src_owner_name', 'src_owner_email', 'src_revision',
               'src_subject', 'src_revision_number', 'change_id', 'src_status', 'src_branch', 'src_hashtags']
    extract_and_save_numbers(commit_lists_web, output_file_excel, headers)

    try:
        # 读取 Excel 文件中的 change 列
        df, changes = read_excel_changes(output_file_excel)
        # 创建新的列来存储查询结果
        df["src_owner_name"] = ""
        df["src_url"] = ""
        df["src_owner_email"] = ""
        df["change_id"] = ""
        df["src_revision"] = ""
        df["src_revision_number"] = ""
        df["src_hashtags"] = ""
        df["src_subject"] = ""
        df["src_status"] = ""
        df["src_branch"] = ""
        total_rows = len(df)
        for index, change in enumerate(changes):
            current_row = index + 1
            print(f"{current_row}/{total_rows}")
            try:
                output_json = download_gerrit_info_analysis(args, change, 'detail')
                url = dump_url + str(change)
                owner_name = ""
                owner_email = ""
                change_id = ""
                current_revision_number = ""
                hashtags = ""
                subject = ""
                status = ""
                branch = ""
                current_revision = ""
                if os.path.exists(output_json):
                    with open(output_json, 'r') as f:
                        data = json.load(f)

                    owner_name = data.get("owner", {}).get("name", "0")
                    owner_email = data.get("owner", {}).get("email", "0")
                    change_id = data.get("change_id", '0')
                    current_revision_number = data.get("current_revision_number", '0')
                    hashtags = ", ".join(data.get("hashtags", []))
                    subject = data.get("subject", '0')
                    status = data.get("status", '0')
                    branch = data.get("branch", '0')
                    # 打印变量
                    print(f"Owner Name: {owner_name}")
                    print(f"Owner Email: {owner_email}")
                    print(f"Change ID: {change_id}")
                    print(f"Revision Number: {current_revision_number}")
                    print(f"Hashtags: {hashtags}")
                    print(f"Subject: {subject}")
                    print(f"Status: {status}")
                    print(f"Branch: {branch}")

                output_json = download_gerrit_info_analysis(args, change, 'revision')
                if os.path.exists(output_json):
                    with open(output_json, 'r') as f:
                        data = json.load(f)
                    item = None
                    current_revision = ''
                    if isinstance(data, list):
                        if len(data) > 1:
                            print("日志: 数据包含超过一个数组")
                        elif data:
                            item = data[0]
                            # print(f"item: {item}")
                    else:
                        item = data
                        # print(f"item: {item}")
                    if item:
                        current_revision = item.get("current_revision", '0')

                # 将提取的信息存储到 DataFrame 中
                df.at[index, "src_url"] = url
                df.at[index, "src_owner_name"] = owner_name
                df.at[index, "src_owner_email"] = owner_email
                df.at[index, "src_revision"] = current_revision
                df.at[index, "change_id"] = change_id
                df.at[index, "src_revision_number"] = current_revision_number
                df.at[index, "src_hashtags"] = hashtags
                df.at[index, "src_subject"] = subject
                df.at[index, "src_status"] = status
                df.at[index, "src_branch"] = branch

            except Exception as e:
                print(f"Error processing change {change}: {e}")

        # 将结果写入新的 Excel 文件
        df.to_excel(output_file_excel, index=False)
        # print(f"Results saved to {output_file_excel}")

    except Exception as e:
        print(f"Error: {e}")


def get_modified_files(repo_path, revision):
    # Change the current working directory to the repo path
    original_cwd = os.getcwd()
    os.chdir(repo_path)

    try:
        # Run the git diff-tree command
        result = subprocess.run(
            ['git', 'diff-tree', '--no-commit-id', '--name-only', '-r', revision],
            capture_output=True,
            text=True,
            check=True
        )
        modified_files = result.stdout.split('\n')
        return modified_files
    finally:
        # Change back to the original working directory
        os.chdir(original_cwd)


def check_modified_files(modified_files):
    for file in modified_files:
        if 'android/abi_gki_' in file:
            return 'symbol'
    return 'code'


def check_commits_type(args, output_file_excel):
    # 读取 Excel 文件
    df = pd.read_excel(output_file_excel)

    # 确保 revision 和 hashtags 列存在
    if 'src_revision' not in df.columns or 'src_hashtags' not in df.columns:
        raise ValueError("Excel 文件缺少 'revision' 或 'hashtags' 列")

    # 创建 type 列并根据 revision 和 hashtags 设置其值
    type_values = []
    for index, row in df.iterrows():
        revision = row['src_revision']
        hashtags = pd.notna(row['src_hashtags'])
        # 使用 fillna 处理空值
        # print(f"Revision: {revision}, Hashtags: {hashtags}")
        if hashtags:
            type_value = 'kmi'
        else:
            modified_files = get_modified_files(args.common, revision)
            type_value = check_modified_files(modified_files)
        # print(type_value)
        type_values.append(type_value)

    df['src_type'] = type_values

    # 保存修改后的 Excel 文件
    df.to_excel(output_file_excel, index=False)


def get_dst_commit_info(args, output_file_excel):
    try:
        # 读取 Excel 文件中的 change 列
        df, change_ids = read_excel_change_ids(output_file_excel)
        df["dst_change"] = ""
        df["dst_url"] = ""
        df["dst_branch"] = ""
        df["dst_status"] = ""

        total_rows = len(df)
        for index, change_id in enumerate(change_ids):
            current_row = index + 1
            print(f"{current_row}/{total_rows}")
            try:
                output_json = download_gerrit_info_analysis(args, change_id, 'changeid')
                item = None
                if os.path.exists(output_json):
                    with open(output_json, 'r') as f:
                        data = json.load(f)

                    # 处理返回的列表
                    if isinstance(data, list):
                        if len(data) > 1:
                            print("日志: 数据包含超过一个数组")
                        elif data:
                            item = data[0]
                            # print(f"item: {item}")
                    else:
                        item = data
                        # print(f"item: {item}")

                    if item:
                        change = item.get("virtual_id_number", '0')
                        url = dump_url + str(change)
                        df.at[index, "dst_change"] = change
                        df.at[index, "dst_url"] = url
                        df.at[index, "dst_branch"] = item.get("branch", '0')
                        df.at[index, "dst_status"] = item.get("status", '0')

            except Exception as e:
                print(f"Error processing change {change_id}: {e}")

        # 将结果写入新的 Excel 文件
        df.to_excel(output_file_excel, index=False)
        # print(f"Results saved to {output_file_excel}")

    except Exception as e:
        print(f"Error: {e}")


def process_excel(file_name, sheet_name=None, committer_email="zuoyonghua@oppo.com"):
    try:
        # 加载 Excel 文件
        wb = load_workbook(file_name, data_only=True)

        # 选择工作表
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.worksheets[0]  # 默认使用第一个工作表

        # 获取表头行
        header_row = sheet[1]
        header_values = [cell.value for cell in header_row]

        # 检查必需的列是否存在
        required_columns = ["Committer_Email", "name", "department"]
        if not all(col in header_values for col in required_columns):
            print(f"Required columns ({', '.join(required_columns)}) not found in the header.")
            return

        # 获取列索引
        committer_email_index = header_values.index("Committer_Email")
        name_index = header_values.index("name")
        department_index = header_values.index("department")

        # 遍历每一行
        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
            if len(row) <= max(committer_email_index, name_index, department_index):
                continue  # 跳过不完整的行

            current_committer_email = row[committer_email_index]
            if current_committer_email == committer_email:
                name_value = row[name_index]
                department_value = row[department_index]
                results.append((name_value, department_value))

        return results

    except FileNotFoundError:
        print(f"File {file_name} not found.")
    except Exception as e:
        print(f"An error occurred: {e}")


def read_excel_email(file_path):
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 提取 change 列的值并转换为字符串
    change_ids = df.get("src_owner_email", []).astype(str)

    return df, change_ids


def get_email_commit_info(args, output_file_excel):
    try:
        # 读取 Excel 文件中的 change 列
        df, emails = read_excel_email(output_file_excel)
        df["src_cn_name"] = ""
        df["src_department"] = ""
        total_rows = len(df)
        for index, email in enumerate(emails):
            current_row = index + 1
            print(f"{current_row}/{total_rows} {email} detail info")
            try:
                results = process_excel(args.config_info, sheet_name=None, committer_email=email)
                if results:
                    for name, department in results:
                        # print(f"Name: {name}, Department: {department}")
                        df.at[index, "src_cn_name"] = name
                        df.at[index, "src_department"] = department
                else:
                    print("No matching entries found.")

            except Exception as e:
                print(f"Error processing change {email}: {e}")

        # 将结果写入新的 Excel 文件
        df.to_excel(output_file_excel, index=False)
        # print(f"Results saved to {output_file_excel}")

    except Exception as e:
        print(f"Error: {e}")


def extract_revision_date_from_file(xml_file_path, path_value):
    # 检查文件是否存在
    if not os.path.exists(xml_file_path):
        print(f"File {xml_file_path} does not exist.")
        return None

    # 读取文件内容
    with open(xml_file_path, 'r') as file:
        xml_content = file.read()

    # 定义正则表达式模式，排除空格的影响
    pattern = re.compile(rf'<project\s+path\s*=\s*"{re.escape(path_value)}"\s+[^>]*revision\s*=\s*"([^"]+)"')

    # 查找匹配的行
    match = pattern.search(xml_content)

    if match:
        revision = match.group(1)
        # 定义日期格式的正则表达式
        date_pattern = re.compile(r'(\d{4}-\d{2})')
        date_match = date_pattern.search(revision)

        if date_match:
            return date_match.group(1)

    return None


def get_nearest_tag_and_date(directory):
    # 检查目录是否存在
    if not os.path.exists(directory):
        print(f"Directory {directory} does not exist.")
        return None

    # 进入指定目录
    original_cwd = os.getcwd()
    os.chdir(directory)

    try:
        # 获取最新的 commit hash
        current_commit = subprocess.check_output(['git', 'rev-parse', 'HEAD']).strip().decode('utf-8')
        print(f"Current commit hash: {current_commit}")

        # 查找最近的 TAG
        try:
            latest_tag = subprocess.check_output(
                ['git', 'describe', '--tags', '--abbrev=0', '--always', current_commit]).strip().decode('utf-8')
        except subprocess.CalledProcessError:
            latest_tag = None

        if not latest_tag or latest_tag == current_commit:
            print("No tags found for current commit.")
            return None

        print(f"Nearest tag found: {latest_tag}")

        # 检查 TAG 是否符合 android15-6.6-2024-12_r2 格式
        tag_pattern = re.compile(r'^android[^-]+-[^-]+-\d{4}-\d{2}_r\d+$')
        if not tag_pattern.match(latest_tag):
            print("Tag does not match the required format android15-6.6-2024-12_r2")
            return None

        # 提取符合 YYYY-MM 格式的日期部分
        print(f"latest_tag: {latest_tag}")
        date_pattern = re.compile(r'(\d{4}-\d{2})')
        date_match = date_pattern.search(latest_tag)

        if date_match:
            date = date_match.group(1)
            print(f"Extracted date: {date}")
            return date
        else:
            print("No valid date found in the tag.")
            return None

    except subprocess.CalledProcessError as e:
        print(f"Error executing git command: {e}")
        return None
    finally:
        # 返回原始目录
        os.chdir(original_cwd)


def dump_gerrit_commits_info(args):
    get_commit_lists_from_file(args, args.commit_lists_web, args.commit_lists_id)
    check_commits_type(args, args.commit_lists_id)
    get_email_commit_info(args, args.commit_lists_id)
    get_dst_commit_info(args, args.commit_lists_id)


def build_ack(args):
    run_repo_command(args)
    max_commits_list_excel = os.path.join(args.log, "max_commits_list.xlsx")
    if args.gerrit_number:
        merged_hash, branch = pull_to_revision(args, max_commits_list_excel)
    else:
        merged_hash = ""
        branch = ""
        print('args.gerrit_number is null ignore pull_to_revision')

    if args.auto:
        merged_hash = update_valid_commits_list_and_push(args, max_commits_list_excel, branch, merged_hash)
    build_kernel_cmd(args, args.bug_number)
    git_format_patch(args, args.common, args.out, merged_hash)
    tar_log = os.path.join(args.log, "tar_log.txt")
    copy_files_to_dist(args.out_abi, file_paths, args.out)

    zip_log = os.path.join(args.out, "log.tar.gz")
    log_out = os.path.join(args.out, "log.txt")

    tar_dir_to_gz(zip_log, args.log, log_out)
    get_build_info(os.path.join(args.out, "BUILD_INFO.txt"))

    manifests_path = os.path.join(args.path, '.repo/manifests/default.xml')
    tag_date = extract_revision_date_from_file(manifests_path, "common")
    if not tag_date:
        tag_date = get_nearest_tag_and_date(args.common)
    if tag_date:
        print(f"Valid date found: {tag_date}")
    else:
        print("No valid date found.")
    branch = extract_android_version(args.branch)
    tmp_name = generate_build_file_name(args, branch, tag_date, "0", args.bug_number)
    tmp_path = os.path.join(args.path, tmp_name)
    create_symlink(args.out, tmp_path)
    current_release = get_current_release(args, tmp_path)
    final_name = generate_build_file_name(args, branch, tag_date, current_release, args.bug_number)
    final_path = os.path.join(args.path, final_name)
    print("final_path:", final_path)
    delete_symlink(tmp_path)
    create_symlink(args.out, final_path)
    final_zip = os.path.join(args.out, final_name + ".tar.gz")
    print("final_zip:", final_zip)
    tar_dir_to_gz(final_zip, args.out_abi, tar_log)
    args.out = final_path
    upload_oem_ki_build(args, final_path, os.path.join(args.log, "upload_file.txt"))
    default_set_gerrit_config_info(args, args.user, args.type, args.config_info)


def main():
    print('\nget input argc/argv')
    print('you can read this link for more help\n')

    args = parse_cmd_args()
    init_work_env(args, args.path, args.out)
    if args.action.lower() == 'dump':
        dump_gerrit_commits_info(args)
    else:
        build_ack(args)


if __name__ == "__main__":
    main()
