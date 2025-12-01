import os
import sys
import subprocess
import pathlib
from datetime import datetime
import openpyxl
from openpyxl import load_workbook, Workbook
import argparse
import common


def get_tools_path(relative_path):
    base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    return full_path


def generate_abi(output_dir):
    generate_abi_gki_aarch64_oplus = "kernel_platform/build/abi/extract_symbols {}/ " \
                                     "--skip-module-grouping --symbol-list " \
                                     "{}/abi_gki_aarch64_oplus " \
                                     "| grep 'Symbol '>> " \
                                     "{}/abi_required_full.txt".format(output_dir, output_dir, output_dir)
    process = subprocess.Popen(generate_abi_gki_aarch64_oplus, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    #print("stand output:{}".format(stdout.decode('utf-8')))
    #print("stand error:{}".format(stderr.decode('utf-8')))


def compare_and_link_kos(kernel_dir, base_file, output_dir):
    # 以 Path 对象制造完整路径
    kernel_dir_path = pathlib.Path(kernel_dir)
    output_dir_path = pathlib.Path(output_dir)

    # 确保输出目录存在
    output_dir_path.mkdir(parents=True, exist_ok=True)

    # 获取所有的 .ko 文件
    ko_files = list(kernel_dir_path.glob('**/*.ko'))

    # 将检索到的文件输出到文件列表中
    list_file = output_dir_path / "current_ko_list.txt"
    list_file.write_text("\n".join(str(file) for file in ko_files))

    # 读取基准文件名
    with open(base_file, 'r') as f:
        base_kos = set(f.read().splitlines())

    # 对比文件并输出新增到文件中
    new_ko_file = output_dir_path / "new_ko_list.txt"
    with new_ko_file.open('w') as f:
        for ko_file in ko_files:

            if ko_file.name not in base_kos:
                # print("ko_file.name:{}".format(ko_file.name))
                f.write(ko_file.name + '\n')
                # 创建软链接
                link_path = output_dir_path / ko_file.name
                if not link_path.exists():
                    link_path.symlink_to(ko_file.resolve())


def extract_first_column(input_file, output_file):
    """
    读取 CSV 文件的每一行，提取每行的第一列，并将其输出到另一个文件中。

    :param input_file: 输入 CSV 文件的路径
    :param output_file: 输出文件的路径，将包含提取的第一列数据
    """
    try:
        # 使用 'with' 语句确保文件正确打开和关闭
        with open(input_file, 'r', encoding='utf-8') as source_file, open(output_file, 'w',
                                                                          encoding='utf-8') as target_file:
            for line in source_file:
                # 使用 strip() 移除尾部的换行符，然后以逗号分割每一行以获取列数据
                columns = line.strip().split(',')
                if columns:  # 确保不处理空行
                    first_column = columns[0]  # 提取第一列
                    target_file.write(first_column + '\n')  # 将第一列数据写入输出文件，并添加换行符
    except IOError as e:
        print("file error: {}".format(e))


def find_common_lines_and_write_sorted(file1_path, file2_path, output_file_path):
    # 使用 Path 对象来操作文件路径
    file1 = pathlib.Path(file1_path)
    file2 = pathlib.Path(file2_path)
    output_file = pathlib.Path(output_file_path)

    # 读取文件内容到集合中
    with open(file1, 'r') as f:
        lines_file1 = set(f.read().splitlines())
    with open(file2, 'r') as f:
        lines_file2 = set(f.read().splitlines())

    # 找到两个文件共有的行
    common_lines = lines_file1.intersection(lines_file2)

    # 对结果进行排序
    sorted_common_lines = sorted(common_lines)

    # 将排序后的共有行写入输出文件
    with open(output_file, 'w') as f:
        f.writelines("\n".join(sorted_common_lines))


def compare_files(file_a, file_b, common_file, only_a_file, only_b_file):
    """
    对比两个文件并输出结果到三个不同的文件中。

    :param file_a: 文件 A 的路径
    :param file_b: 文件 B 的路径
    :param common_file: 公共行将被写入的文件路径
    :param only_a_file: 只存在于文件 A 的行将被写入的文件路径
    :param only_b_file: 只存在于文件 B 的行将被写入的文件路径
    """
    with open(file_a, 'r', encoding='utf-8') as a, open(file_b, 'r', encoding='utf-8') as b:
        # 读取两个文件的内容到集合中，自动去除重复的行
        lines_a = set(a.readlines())
        lines_b = set(b.readlines())

    # 找到共同的行和独有的行
    common_lines = lines_a.intersection(lines_b)
    only_a_lines = lines_a - lines_b
    only_b_lines = lines_b - lines_a

    # 将结果写入到三个不同的文件，并进行排序
    with open(common_file, 'w', encoding='utf-8') as cf:
        cf.writelines(sorted(common_lines))

    with open(only_a_file, 'w', encoding='utf-8') as oaf:
        oaf.writelines(sorted(only_a_lines))

    with open(only_b_file, 'w', encoding='utf-8') as obf:
        obf.writelines(sorted(only_b_lines))


def txt_to_excel(txt_file, excel_file):
    """
    将一个文本文件转换为Excel文件，文本内容以空格分割。

    :param txt_file: 输入的文本文件路径
    :param excel_file: 输出的Excel文件路径
    """
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(txt_file, 'r', encoding='utf-8') as file:
        for row_index, line in enumerate(file, start=1):
            # 使用空格对每一行进行分割
            columns = line.strip().split(" ")
            for col_index, cell_value in enumerate(columns, start=1):
                # 将分割后的数据写入Excel工作表
                ws.cell(row=row_index, column=col_index).value = cell_value

    # 保存Excel工作簿
    wb.save(excel_file)
    # print("{} has been converted to {}".format(txt_file,excel_file))


def compare_and_update_excel(common_file, ko_owner_file):
    # 打开 common.xlsx 文件
    common_wb = load_workbook(common_file)
    common_ws = common_wb.active

    # 打开 ko_owner_list.xlsx 文件
    ko_owner_wb = load_workbook(ko_owner_file)
    ko_owner_ws = ko_owner_wb.active

    # 对 common.xlsx 的每一行进行检查和更新
    for row in common_ws.iter_rows(min_row=1, min_col=2, max_col=2):
        common_cell = row[0]
        # 在 ko_owner_list.xlsx 中查找对应的行
        for ko_row in ko_owner_ws.iter_rows(min_row=1):
            ko_owner_cell = ko_row[0]
            if ko_owner_cell.value == common_cell.value:
                # 如果找到匹配，则更新 common.xlsx 的相应单元格
                common_ws.cell(row=common_cell.row, column=3, value=ko_row[1].value)
                common_ws.cell(row=common_cell.row, column=4, value=ko_row[2].value)
                common_ws.cell(row=common_cell.row, column=5, value=ko_row[3].value)
                break

    # 保存更新后的 common.xlsx 文件
    # common_wb.save('updated_common.xlsx')
    common_wb.save(common_file)


def merge_excel_sheets(input_files, output_file):
    # 创建一个新的工作簿
    output_wb = Workbook()

    # 移除默认创建的空sheet
    output_wb.remove(output_wb.active)

    for file_path in input_files:
        # 加载每个Excel文件
        input_wb = load_workbook(file_path)
        # 获取第一个工作表
        input_ws = input_wb.active
        # 创建一个新的工作表，名称与输入文件相同
        sheet_name = file_path.split('/')[-1].replace('.xlsx', '')
        output_ws = output_wb.create_sheet(title=sheet_name)

        # 复制单元格数据到新的工作表
        for row in input_ws.iter_rows():
            for cell in row:
                output_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

    # 保存这个新的Excel文件
    output_wb.save(output_file)


def rename_dlkm_files(raw_img_path):
    output_dir = os.path.dirname(raw_img_path)
    file_renames = {
        "system_dlkm_a.img": "system_dlkm.img",
        "vendor_dlkm_a.img": "vendor_dlkm.img",
        "odm_dlkm_a.img": "odm_dlkm.img"
    }
    common.rename_files(output_dir, file_renames)


# Example usage
def unpack_dlkm_images(paths):
    dlkm_file_names = ['system_dlkm.img', 'vendor_dlkm.img', 'odm_dlkm.img']
    # 遍历路径列表
    for path in paths:
        if os.path.exists(path):
            # 遍历文件名列表
            for file_name in dlkm_file_names:
                dlkm_img = os.path.join(path, file_name)
                if os.path.exists(dlkm_img):
                    common.unpack_dlkm_image(dlkm_img)
    return


def unpack_vendor_boot_imgs(paths):
    local_file_names = ['vendor_boot.img']
    # 遍历路径列表
    for path in paths:
        if os.path.exists(path):
            # 遍历文件名列表
            for file_name in local_file_names:
                boot_img = os.path.join(path, file_name)
                if os.path.exists(boot_img):
                    common.unpack_vendor_boot_img(boot_img)


def parse_cmd_args():
    parser = argparse.ArgumentParser(description="abi check for oki gki ogki")

    parser.add_argument('-o', '--out', type=str, help='Kernel tmp out dir (default: tools/out)',
                        default=get_tools_path("out"))
    parser.add_argument('--qcom_master_path', type=str, help='', default=get_tools_path("out/qcom_master"))
    parser.add_argument('--qcom_do_path', type=str, help='', default=get_tools_path("out/qcom_do"))
    parser.add_argument('--mtk_master_path', type=str, help='', default=get_tools_path("out/mtk_master"))
    parser.add_argument('--mtk_do_path', type=str, help='', default=get_tools_path("out/mtk_do"))
    parser.add_argument('--qcom_master_url', type=str, help='xxx/IMAGES/', default="")
    parser.add_argument('--qcom_do_url', type=str, help='xxx/do/', default="")
    parser.add_argument('--mtk_master_url', type=str, help='xxx/IMAGES/', default="")
    parser.add_argument('--mtk_do_url', type=str, help='xxx/do/', default="")
    parser.add_argument('--type', type=str, help='', default="unpack")

    args = parser.parse_args()
    paths = [
        pathlib.Path(args.qcom_master_path),
        pathlib.Path(args.mtk_master_path),
        pathlib.Path(args.qcom_do_path),
        pathlib.Path(args.mtk_do_path)
        # 可以在这里动态添加其他路径
    ]
    args.paths = paths
    current_path = os.getcwd()
    args.out_put = os.path.join(current_path, "out/oplus")
    args.log_dir = os.path.join(current_path, 'LOGDIR/abi/')
    args.exclude_file = ['*.ko', 'vmlinux']
    for key, value in vars(args).items():
        print("{}: {}".format(key, value))
    return args


def download_prebuilt_images(args):
    file_names = ['vendor_boot.img', 'vendor_dlkm.img', 'system_dlkm.img', 'odm_dlkm.img', 'super.img']
    download_configs = [
        {
            "path": args.qcom_master_path,
            "url": args.qcom_master_url
        },
        {
            "path": args.mtk_master_path,
            "url": args.mtk_master_url
        },
        {
            "path": args.qcom_do_path,
            "url": args.qcom_do_url
        },
        {
            "path": args.mtk_do_path,
            "url": args.mtk_do_url
        }
    ]
    # 使用 for 循环遍历每个配置并下载文件
    for config in download_configs:
        path = config["path"]
        url = config["url"]
        if url.startswith('http'):
            common.download_files_from_server(path, file_names, url)
        elif url:
            common.copy_files_form_src_to_dst(path, file_names, url)
        else:
            print("Invalid URL {}".format(url))

    return


def get_dlkm_img_from_super(paths):
    sparse_file_names = ['super.img']
    # 遍历路径列表
    for path in paths:
        if os.path.exists(path):
            # print(f"Checking path: {path}")
            # 遍历文件名列表
            for file_name in sparse_file_names:
                super_img = os.path.join(path, file_name)
                if os.path.exists(super_img):
                    print(f"Found file: {super_img}")
                    common.unpack_superimg(super_img)
                    rename_dlkm_files(super_img)

    return


def find_and_write_files(src_dir_list, file_pattern, output_file):
    ko_files = set()

    for src_dir in src_dir_list:
        src_path = pathlib.Path(src_dir)
        for file in src_path.rglob(file_pattern):
            ko_files.add(file.name)

    sorted_ko_files = sorted(ko_files)
    # 确保输出文件的目录存在
    output_path = pathlib.Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, 'w') as f:
        for file_name in sorted_ko_files:
            f.write(file_name + '\n')


def get_all_ko(args):
    local_file_names = ['system_dlkm', 'system_dlkm_mnt', 'vendor_boot', 'vendor_boot_mnt', 'vendor_dlkm',
                        'vendor_dlkm_mnt', 'odm_dlkm', 'odm_dlkm_mnt']
    # 遍历路径列表
    for path in args.paths:
        target_dir = os.path.join(path, "all_ko/")
        output_file = os.path.join(path, "all_ko/ko_files.txt")
        if os.path.exists(path):
            print(f"Checking path: {path}")
            # 遍历文件名列表
            for file_name in local_file_names:
                file_path = os.path.join(path, file_name)
                if os.path.exists(file_path):
                    # print(f"Found file: {file_path}")
                    local_path = [pathlib.Path(file_path)]
                    # print(f"Found file: {local_path}")
                    common.src_to_dst_from_directories(local_path, "*.ko", target_dir, 'copy')

        src_dirs = [pathlib.Path(target_dir)]
        find_and_write_files(src_dirs, "*.ko", output_file)


def prebuilt_image_download(args):
    download_prebuilt_images(args)
    get_dlkm_img_from_super(args.paths)
    unpack_dlkm_images(args.paths)
    unpack_vendor_boot_imgs(args.paths)
    get_all_ko(args)


def main():
    start_time = datetime.now()
    print("Begin time:", start_time.strftime("%Y-%m-%d %H:%M:%S"))
    print("get symbol list start ...")

    args = parse_cmd_args()
    common.set_common_build_env()
    prebuilt_image_download(args)
    if args.type== "unpack":
        print("unpack finish")
        sys.exit(0)

    print("[1] delete_all_exists_file...")
    common.delete_all_exists_file(args.out_put)
    print("[2] create_symlinks for vmlinux to qcom all...")

    qcom_master_src = pathlib.Path(args.qcom_master_path) / "all_ko"
    qcom_do_src = pathlib.Path(args.qcom_do_path) / "all_ko"
    qcom_master_dst = pathlib.Path(args.out_put) / "qcom/master"
    qcom_oplus_dst = pathlib.Path(args.out_put) / "qcom/oplus"

    mtk_master_src = pathlib.Path(args.mtk_master_path) / "all_ko"
    mtk_do_src = pathlib.Path(args.mtk_do_path) / "all_ko"
    mtk_master_dst = pathlib.Path(args.out_put) / "mtk/master"
    mtk_oplus_dst = pathlib.Path(args.out_put) / "mtk/oplus"

    mtk_master_src_list = [mtk_master_src]
    qcom_master_src_list = [qcom_master_src]

    src_dst_qcom_master = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/qcom/master/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/qcom/master/vmlinux'),
    ]
    src_dst_qcom_oplus = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/qcom/oplus/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/qcom/oplus/vmlinux'),
    ]
    src_dst_mtk_master = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/mtk/master/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/mtk/master/vmlinux'),
    ]
    src_dst_mtk_oplus = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/mtk/oplus/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/mtk/oplus/vmlinux'),
    ]
    qcom_master_file_path = "{}/abi_required_full.txt".format(qcom_master_dst)
    qcom_master_output_list = "{}/abi_required_symbol.txt".format(qcom_master_dst)
    qcom_master_output_ko = "{}/abi_required_ko.txt".format(qcom_master_dst)
    qcom_master_output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(qcom_master_dst)

    qcom_oplus_file_path = "{}/abi_required_full.txt".format(qcom_oplus_dst)
    qcom_oplus_output_list = "{}/abi_required_symbol.txt".format(qcom_oplus_dst)
    qcom_oplus_output_ko = "{}/abi_required_ko.txt".format(qcom_oplus_dst)
    qcom_oplus_output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(qcom_oplus_dst)

    mtk_master_file_path = "{}/abi_required_full.txt".format(mtk_master_dst)
    mtk_master_output_list = "{}/abi_required_symbol.txt".format(mtk_master_dst)
    mtk_master_output_ko = "{}/abi_required_ko.txt".format(mtk_master_dst)
    mtk_master_output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(mtk_master_dst)

    mtk_oplus_file_path = "{}/abi_required_full.txt".format(mtk_oplus_dst)
    mtk_oplus_output_list = "{}/abi_required_symbol.txt".format(mtk_oplus_dst)
    mtk_oplus_output_ko = "{}/abi_required_ko.txt".format(mtk_oplus_dst)
    mtk_oplus_output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(mtk_oplus_dst)

    common.create_symlinks_from_pairs(src_dst_qcom_master)

    common.src_to_dst_from_directories(qcom_master_src_list, "*.ko", qcom_master_dst)

    print("[4] generate qcom all abi...")
    generate_abi(qcom_master_dst)

    print("[5] get qcom all symbols and ko list...")
    common.get_symbols_and_ko_list(qcom_master_file_path, qcom_master_output_list, qcom_master_output_ko,
                                   qcom_master_output_symbol_ko)

    print("[6] create_symlinks for qcom oplus vmlinux ...")
    common.create_symlinks_from_pairs(src_dst_qcom_oplus)

    compare_and_link_kos(
        kernel_dir=qcom_master_src,
        base_file=qcom_do_src / 'ko_files.txt',
        output_dir=qcom_oplus_dst
    )
    print("[7] generate qcom oplus abi...")
    generate_abi(qcom_oplus_dst)
    print("[8] get qcom oplus symbols and ko list...")
    common.get_symbols_and_ko_list(qcom_oplus_file_path, qcom_oplus_output_list,
                                   qcom_oplus_output_ko, qcom_oplus_output_symbol_ko)

    print("[9] get qcom platform oplus require symbols...")
    find_common_lines_and_write_sorted(
        'out/oplus/qcom/oplus/abi_required_symbol_ko.txt',
        'out/oplus/qcom/master/abi_required_symbol_ko.txt',
        'out/oplus/qcom/output_sorted_common_lines.txt'
    )

    # mtk platform
    print("\nmtk platform\n")
    print("[1] null")
    # example: Predefining source and target files

    common.create_symlinks_from_pairs(src_dst_mtk_master)
    print("[2] create_symlinks for vmlinux to mtk all...")

    common.src_to_dst_from_directories(mtk_master_src_list, "*.ko", mtk_master_dst)
    print("[3] create symlinks for mtk all ko...")

    print("[4] generate mtk all abi...")
    generate_abi(mtk_master_dst)

    print("[5] get mtk all symbols and ko list...")
    common.get_symbols_and_ko_list(mtk_master_file_path, mtk_master_output_list,
                                   mtk_master_output_ko, mtk_master_output_symbol_ko)

    print("[6] create_symlinks for mtk oplus vmlinux ...")
    common.create_symlinks_from_pairs(src_dst_mtk_oplus)

    compare_and_link_kos(
        kernel_dir=mtk_master_src,
        base_file=mtk_do_src / 'ko_files.txt',
        output_dir=mtk_oplus_dst
    )

    print("[7] generate mtk oplus abi...")
    generate_abi(mtk_oplus_dst)

    print("[8] get mtk oplus symbols and ko list...")
    common.get_symbols_and_ko_list(mtk_oplus_file_path, mtk_oplus_output_list,
                                   mtk_oplus_output_ko, mtk_oplus_output_symbol_ko)

    print("[9] get mtk platform oplus require symbols...")
    find_common_lines_and_write_sorted(
        'out/oplus/mtk/oplus/abi_required_symbol_ko.txt',
        'out/oplus/mtk/master/abi_required_symbol_ko.txt',
        'out/oplus/mtk/output_sorted_common_lines.txt'
    )

    print("get all need symbol symbols...")
    compare_files('out/oplus/qcom/output_sorted_common_lines.txt',
                  'out/oplus/mtk/output_sorted_common_lines.txt',
                  'out/oplus/common.txt',
                  'out/oplus/qcom_only.txt',
                  'out/oplus/mtk_only.txt')

    txt_to_excel('out/oplus/common.txt', 'out/oplus/common.xlsx')
    txt_to_excel('out/oplus/qcom_only.txt', 'out/oplus/qcom_only.xlsx')
    txt_to_excel('out/oplus/mtk_only.txt', 'out/oplus/mtk_only.xlsx')

    compare_and_update_excel('out/oplus/common.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    compare_and_update_excel('out/oplus/qcom_only.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    compare_and_update_excel('out/oplus/mtk_only.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    input_files = [
        'out/oplus/common.xlsx',
        'out/oplus/qcom_only.xlsx',
        'out/oplus/mtk_only.xlsx'
    ]
    output_file = 'out/oplus/android16-6.12.xlsx'
    merge_excel_sheets(input_files, output_file)

    print("get symbol list end ...")
    end_time = datetime.now()
    print("End time:", end_time.strftime("%Y-%m-%d %H:%M:%S"))
    elapsed_time = end_time - start_time
    print("Total time:", common.format_duration(elapsed_time))


if __name__ == "__main__":
    main()
