import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def csv_to_xlsx(csv_filename, xlsx_filename):
    """
    Convert a CSV file to an Excel file.

    :param csv_filename: The name of the CSV file to read from.
    :param xlsx_filename: The name of the Excel file to write to.
    """
    # Check if the CSV file exists
    if not os.path.isfile(csv_filename):
        print(f"CSV file {csv_filename} does not exist.")
        return

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Read the CSV file and write its contents to the Excel file
    with open(csv_filename, mode='r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file)
        for row in csv_reader:
            ws.append(row)

    # Save the Excel file
    wb.save(xlsx_filename)
    print(f"CSV file {csv_filename} has been converted to Excel file {xlsx_filename}.")


def process_with_openpyxl(input_file, output_file):

    # 加载Excel文件
    workbook = load_workbook(filename=input_file)

    # 假设数据在第一个工作表中
    sheet = workbook.active

    # 获取列标题
    headers = [cell.value for cell in sheet[1]]

    # 检查所需的列是否存在
    required_columns = ['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol', 'name',
                        'config_userid', 'project']
    if not all(column in headers for column in required_columns):
        print(f"文件中缺少所需的列: {required_columns}")
        return

    # 获取所需列的索引
    gerrit_projects_index = headers.index('gerrit_projects')
    gerrit_branch_index = headers.index('gerrit_branch')
    release_version_index = headers.index('release_version')
    release_eol_index = headers.index('release_eol')
    name_index = headers.index('name')
    config_userid_index = headers.index('config_userid')
    project_index = headers.index('project')

    # 创建一个新的工作簿和工作表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # 写入新的列标题
    new_sheet.append(['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol', 'config_userid', 'message'])

    # 使用字典来存储分组后的数据
    grouped_data = {}

    # 遍历每一行并处理数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        gerrit_projects = row[gerrit_projects_index]
        gerrit_branch = row[gerrit_branch_index]
        release_version = row[release_version_index]
        release_eol = row[release_eol_index]
        name = row[name_index]
        config_userid = row[config_userid_index]
        project = str(row[project_index])

        # 合并 'name' 和 'project' 列的内容到 'message' 列
        message = f"{name}:{project}"

        # 创建分组键
        key = (gerrit_projects, gerrit_branch, release_version, release_eol, config_userid)

        # 将消息添加到分组数据中
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(message)

    # 将分组后的数据写入新的工作表
    for key, messages in grouped_data.items():
        new_sheet.append([*key, '\n'.join(messages)])

    # 保存为新的CSV文件
    new_workbook.save(output_file)
    print(f"file save to {output_file}")


def process_with_pandas(input_file, output_file):
    import pandas as pd

    try:
        # 读取CSV文件
        df = pd.read_csv(input_file)

        # 选择需要的列，并确保创建一个新的DataFrame
        required_columns = ['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol', 'name',
                            'config_userid', 'project']
        if not all(column in df.columns for column in required_columns):
            raise ValueError(f"文件中缺少所需的列: {required_columns}")

        selected_columns = df[required_columns].copy()

        # 将 'project' 列转换为字符串类型
        selected_columns.loc[:, 'project'] = selected_columns['project'].astype(str)

        # 合并 'name' 和 'project' 列的内容到 'message' 列
        selected_columns.loc[:, 'message'] = selected_columns['name'] + ':' + selected_columns['project']

        # 按指定列进行分组，并合并 'message' 列的内容
        grouped_df = selected_columns.groupby(['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol',
                                               'config_userid'])['message'].apply(lambda x: '\n'.join(x)).reset_index()

        # 保存为新的CSV文件
        grouped_df.to_csv(output_file, index=False)
        print(f"文件已优化并保存为 {output_file}")
    except Exception as e:
        raise e


def process_with_csv_module(input_file, output_file):
    try:
        # 读取CSV文件
        with open(input_file, mode='r', newline='', encoding='utf-8') as infile:
            reader = csv.DictReader(infile)
            data = list(reader)

        # 检查所需的列是否存在
        required_columns = ['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol', 'name',
                            'config_userid', 'project']
        if not all(column in reader.fieldnames for column in required_columns):
            raise ValueError(f"文件中缺少所需的列: {required_columns}")

        # 创建一个新的列表来存储处理后的数据
        processed_data = []

        # 遍历每一行并处理数据
        for row in data:
            row['project'] = str(row['project'])
            row['message'] = f"{row['name']}:{row['project']}"
            processed_data.append(row)

        # 使用字典来存储分组后的数据
        grouped_data = {}

        for row in processed_data:
            key = (row['gerrit_projects'], row['gerrit_branch'], row['release_version'], row['release_eol'],
                   row['config_userid'])
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(row['message'])

        # 准备写入新的CSV文件的数据
        new_data = []
        for key, messages in grouped_data.items():
            new_row = {
                'gerrit_projects': key[0],
                'gerrit_branch': key[1],
                'release_version': key[2],
                'release_eol': key[3],
                'config_userid': key[4],
                'message': '\n'.join(messages)
            }
            new_data.append(new_row)

        # 写入新的CSV文件
        with open(output_file, mode='w', newline='', encoding='utf-8') as outfile:
            fieldnames = ['gerrit_projects', 'gerrit_branch', 'release_version', 'release_eol', 'config_userid',
                          'message']
            writer = csv.DictWriter(outfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(new_data)

        print(f"file save to {output_file}")
    except Exception as e:
        print(f"使用 csv 模块处理文件时发生错误: {e}")

def optimize_and_save_csv(input_file, output_file):
    try:
        # 尝试使用 pandas 处理 CSV 文件
        process_with_pandas(input_file, output_file)
    except Exception as e:
        # 检查文件扩展名是否为 .csv
        if input_file.lower().endswith('.csv'):
            process_with_csv_module(input_file, output_file)
        elif input_file.lower().endswith('.xlsx'):
            process_with_openpyxl(input_file, output_file)

def ack_csv_to_xlsx(args, csv_file, xlsx_file):
    csv_to_xlsx(csv_file, xlsx_file)

def ack_optimize_and_save_csv(args, csv_file, optimized_csv_file):
    optimize_and_save_csv(csv_file, optimized_csv_file)


