import os
import sys
import requests
import subprocess
import shutil
from datetime import datetime, timedelta
import csv
import argparse
import re
import hashlib
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
import json
from pathlib import Path
from urllib.parse import urljoin
# import ack_tt
import ack_dump
import ack_info
import common
import fnmatch

kernel_info = []

def get_tools_path(relative_path):
    base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    return full_path


def get_current_date():
    today = datetime.today()
    formatted_date = today.strftime("%Y-%m-%d")
    return formatted_date


def parse_cmd_args():
    parser = argparse.ArgumentParser(description="ACK LTS OGKI/GKI/OKI Check")
    parser.add_argument('-b', '--bootimg', type=str, help='Path to boot image (default: boot.img)', default='boot.img')
    parser.add_argument('-k', '--kernelinfo', type=str,
                        help='Kernel version info file path (default: out/kernel_version.txt)',
                        default=get_tools_path("out/kernel_version.txt"))
    parser.add_argument('-o', '--out', type=str, help='Kernel tmp out dir (default: tools/out)',
                        default=get_tools_path("out"))
    parser.add_argument('-z', '--lz4', type=str, help='lz4 process tool (default: tools/lz4)',
                        default=get_tools_path("lz4"))
    parser.add_argument('-n', '--unpack_bootimg', type=str, help='Boot unpack tool (default: tools/unpack_bootimg)',
                        default=get_tools_path("unpack_bootimg"))
    parser.add_argument('-a', '--approved', type=str,
                        help='Approved OGKI builds (default: out/tools/kernel_configs/approved-ogki-builds.xml)',
                        default=get_tools_path("out/kernel_configs/approved-ogki-builds.xml"))
    parser.add_argument('-l', '--lifetimes', type=str,
                        help='Kernel lifetimes (default: out/tools/kernel_configs/kernel-lifetimes.xml)',
                        default=get_tools_path("out/kernel_configs/kernel-lifetimes.xml"))
    parser.add_argument('--ack_config', type=str,
                        help='Google default kernel config dir (default:out/ tools/kernel_configs)',
                        default=get_tools_path("out/kernel_configs"))
    parser.add_argument('--out_bootimg', type=str, help='Output dir for boot.img (default: out/boot.img)', default='')
    parser.add_argument('--out_unpack', type=str, help='Output dir for unpacked files (default: out/unpack)',
                        default='')
    parser.add_argument('--out_kernel', type=str, help='Output dir for kernel (default: out/unpack/kernel)', default='')
    parser.add_argument('--build_json', type=str, help='Build info from JSON (default: tools/out/build.json)',
                        default=get_tools_path("out/build.json"))
    parser.add_argument('--ack_all_info', type=str, help='ACK info CSV (default: out/tools/ack_all_info.csv)',
                        default=get_tools_path("out/ack_all_info.csv"))
    parser.add_argument('--prj_cfg_info', type=str,
                        help='Project config info CSV (default: out/tools/prj_cfg_info.csv)',
                        default=get_tools_path("out/prj_cfg_info.csv"))
    parser.add_argument('--ogki_prj_approve', type=str,
                        help='ogki config info CSV (default: out/tools/ogki_prj_approve.csv)',
                        default=get_tools_path("out/ogki_prj_approve.csv"))
    parser.add_argument('--tt_message_info', type=str,
                        help='Project config info CSV (default: out/tools/tt_message_info)',
                        default=get_tools_path("out/tt_message_info.csv"))
    parser.add_argument('--current_date', type=str, help='Current date (default: today)', default=get_current_date())
    parser.add_argument('--type', type=str, help='Build send check type (default: build)', default="build")
    parser.add_argument('--project', type=str, help='ACK LTS check project ID (default: 0)', default='0')
    parser.add_argument('--branch', type=str, help='ACK gerrit branch (default: 0)', default='0')
    parser.add_argument('--config_userid', type=str, help='Config user ID (default: 0)', default='0')
    parser.add_argument('--ignore_date_start', type=str, help='Ignore date start (default: 0)', default='0')
    parser.add_argument('--ignore_date_end', type=str, help='Ignore date end (default: 0)', default='0')
    parser.add_argument('--warning_delta', type=str, help='Warning delta days (default: 90)', default='90')
    parser.add_argument('--error_delta', type=str, help='Error delta days (default: 60)', default='60')
    parser.add_argument('--warning_duty', type=str, help='Warning duty days (default: 7)', default='14')
    parser.add_argument('--error_duty', type=str, help='Error duty days (default: 1)', default='7')
    parser.add_argument('--enable_error', type=str, help='Enable error checks (default: 1)', default='1')

    args = parser.parse_args()

    args.out_bootimg = os.path.join(args.out, "boot.img")
    args.out_unpack = os.path.join(args.out, "unpack")
    args.out_kernel = os.path.join(args.out_unpack, "kernel")

    is_valid, args.current_date = is_valid_date(args.current_date)
    if not is_valid:
        print(f"is_valid  {is_valid} return")
        return

    for key, value in vars(args).items():
        print(f"{key}: {value}")

    if not os.path.exists(args.out):
        os.makedirs(args.out)
    current_path = os.getcwd()
    print(f"current_date: {args.current_date}")
    print(f"current_path: {current_path}\n")
    return args

def ensure_file_exists(file_path):
    # 获取文件所在的目录
    directory = os.path.dirname(file_path)
    # 如果目录不存在，则创建目录
    if not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
        print("Created directory: {}".format(directory))
    # 创建文件（如果文件不存在）
    if not os.path.exists(file_path):
        with open(file_path, 'w') as file:
            pass  # 创建空文件
        print("Created file: {}".format(file_path))

def check_file_exists(args):
    file_vars = [args.kernelinfo]
    for file_path in file_vars:
        ensure_file_exists(file_path)

def get_ack_tt_cmd_source_mode(args):
    return True
    """
    if ack_tt.env != "release":
        return True
        # return False
    else:
        return False
    """

def send_user_message_cmd(args, users, message):
    ack_tt_tools = get_tools_path("ack_tt")
    # 检查 ack_tt 文件是否存在
    if not os.path.exists(ack_tt_tools):
        print("Error: {} not found.".format(ack_tt_tools))
        return
    if get_ack_tt_cmd_source_mode(args):

        command = [
            "python3", get_tools_path("ack_tt.py"),
            "--cmd", "send",
            "--users", users,
            "--message", message
        ]

    else:
        command = [
            ack_tt_tools,
            "--cmd", "send",
            "--users", users,
            "--message", message
        ]

    print('\ncommand', ' '.join(command))
    print('\n')

    try:
        result = subprocess.run(command, shell=False, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        output = result.stdout.decode('utf-8').strip()  # 去除首尾空白字符

    except subprocess.CalledProcessError as e:
        print("run command error： {}".format(e))
    return


def get_prj_user_number_cmd(args, project):
    default_data = {"CREATE_USER_NO": ["0"], "SOFTWARE_PM": ["0"], "CMO": ["0"], "PROJECT_CODE": "0", "PROJECT_NAME": "0"}
    ack_tt_tools = get_tools_path("ack_tt")
    # 检查 ack_tt 文件是否存在
    if not os.path.exists(ack_tt_tools):
        print("Error: {} not found.".format(ack_tt_tools))
        return
    if get_ack_tt_cmd_source_mode(args):
        command = [
            "python3", get_tools_path("ack_tt.py"),
            "--cmd", "get",
            "--project", project
        ]
    else:
        command = [
            ack_tt_tools,
            "--cmd", "get",
            "--project", project
        ]
    print('\ncommand', ' '.join(command))

    print('\n')
    try:
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        output = result.stdout.strip()  # 去除首尾空白字符
        # print("output", output)
        # 逐行检查并解析 JSON
        lines = output.splitlines()
        for line in lines:
            line = line.strip()
            if not line:
                continue  # 跳过空行
            try:
                data_dict = json.loads(line)
                print("data_dict", data_dict)
                return data_dict
            except json.JSONDecodeError:
                # print(f"Skipping non-JSON line: {line}")
                pass
        print("No valid JSON found in output")
        return default_data
    except subprocess.CalledProcessError as e:
        print("Command execution failed: {}".format(e))
        return default_data
    except Exception as e:
        print("Unexpected exception occurred: {}".format(e))
        return default_data


def download_compile_ini(bootimg, out):
    try:
        response = requests.get(urljoin(bootimg, 'compile.ini'), stream=True)
    except ImportError:
        print("SSL module is not available. Can't connect to HTTPS URL.")
        return None
    except Exception as e:
        print('An unexpected error occurred:', str(e))
        return None
    if response.status_code != 200:
        print("Server can't connect, please set boot.img and try again")
        return None

    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, 'compile.ini'), 'wb') as f:
        f.write(response.content)
    return response.text.replace("\r\n", "\n").replace("\r", "\n")


def download_boot_img_from_direct_link(direct_link, out, file='boot.img'):
    try:
        boot_img_response = requests.get(direct_link, stream=True)
    except ImportError:
        print("SSL module is not available. Can't connect to HTTPS URL.")
        return None
    except Exception as e:
        print('An unexpected error occurred:', str(e))
        return None
    if boot_img_response.status_code != 200:
        print('Could not reach {}, status code: {}. Skipping download.'.format(direct_link,
                                                                               boot_img_response.status_code))
        return None

    print('download from: ', direct_link, 'to', out, 'file name', file)
    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, file), 'wb') as f:
        for chunk in boot_img_response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)


def download_prebuild_image(bootimg, out):
    if not bootimg:
        bootimg = input("Your server address: ")

    if not bootimg.endswith('/'):
        bootimg += '/'

    compile_ini_text = download_compile_ini(bootimg, out)
    if not compile_ini_text:
        return

    ofp_folder_lines = [line for line in compile_ini_text.split("\n") if "ofp_folder" in line]
    if not ofp_folder_lines:
        print("Didn't find 'ofp_folder' in compile.ini")
        return

    ofp = [line.split('=')[1].strip() for line in compile_ini_text.split('\n') if line.startswith("ofp_folder")][0]
    boot_img_link = urljoin(bootimg, "{}/IMAGES/boot.img".format(ofp))
    print('ofp:', ofp)
    download_boot_img_from_direct_link(boot_img_link, out)
    boot_img = os.path.join(out, "boot.img")
    if os.path.exists(boot_img):
        build_json = urljoin(bootimg, "build.json")
        download_boot_img_from_direct_link(build_json, out, 'build.json')

        manifest = urljoin(bootimg, "XML_INFO/merge_component_manifests.xml")
        download_boot_img_from_direct_link(manifest, out, 'merge_component_manifests.xml')

    else:
        print("{} File does not exist ignore download build.json merge_component_manifests.xml".format(boot_img))


def copy_boot_img_from_local_path(local_path, out):
    print('copy from:', local_path, 'to', out)
    if not os.path.exists(local_path):
        print("{} File does not exist.ignore".format(local_path))
        return
    os.makedirs(out, exist_ok=True)
    try:
        with open(local_path, 'rb') as src_file:
            with open(os.path.join(out, 'boot.img'), 'wb') as dst_file:
                dst_file.write(src_file.read())
    except FileNotFoundError:
        print("Cannot find file: {}".format(local_path))


def handle_boot_img_path(boot_img_path, out):
    if boot_img_path.startswith('http') and not boot_img_path.endswith('boot.img'):
        download_prebuild_image(boot_img_path, out=out)
    elif boot_img_path.startswith('http') and boot_img_path.endswith('boot.img'):
        download_boot_img_from_direct_link(boot_img_path, out)
    elif not boot_img_path.startswith('http'):
        copy_boot_img_from_local_path(boot_img_path, out)
    else:
        print("Invalid boot image path.")


def parse_version(version_str):
    if not version_str:
        print("Version string is empty.")
        return ""

    match = re.search(r'(\d+)\.(\d+)\.(\d+)', version_str)
    if not match:
        print("Invalid version format.")
        return ""

    current_kernel_version = match.group(0)
    current_version = int(match.group(1))
    current_patchlevel = int(match.group(2))
    current_sublevel = int(match.group(3))

    if "-android" in version_str:
        main_version, android_version = version_str.split("-android")
        android_version_parts = android_version.split("-")

        launchversion = int(android_version_parts[0])
        kmi_generation = int(android_version_parts[1])

        commit = 'UNKNOWN'
        match = re.search(r'-g([^\\-]*)(?:-|$)', android_version)
        if match:
            commit = match.group(1)
    else:
        launchversion = 0
        kmi_generation = 0
        commit = 'UNKNOWN'

    return {
        "CURRENT_KERNEL_VERSION": current_kernel_version,
        "CURRENT_VERSION": current_version,
        "CURRENT_PATCHLEVEL": current_patchlevel,
        "CURRENT_SUBLEVEL": current_sublevel,
        "LAUNCHVERSION": launchversion,
        "KMI_GENERATION": kmi_generation,
        "COMMIT": commit
    }


def print_version_info(version_info):
    if not version_info:
        print("version_info string is empty.")
        return
    print("CURRENT_KERNEL_VERSION:", version_info["CURRENT_KERNEL_VERSION"])
    print("CURRENT_VERSION:", version_info["CURRENT_VERSION"])
    print("CURRENT_PATCHLEVEL:", version_info["CURRENT_PATCHLEVEL"])
    print("CURRENT_SUBLEVEL:", version_info["CURRENT_SUBLEVEL"])
    print("LAUNCHVERSION:", version_info["LAUNCHVERSION"])
    print("KMI_GENERATION:", version_info["KMI_GENERATION"])
    print("COMMIT:", version_info["COMMIT"])
    update_kernel_info("CURRENT_KERNEL_VERSION:" + version_info["CURRENT_KERNEL_VERSION"])
    update_kernel_info("CURRENT_VERSION:" + str(version_info["CURRENT_VERSION"]))
    update_kernel_info("CURRENT_PATCHLEVEL:" + str(version_info["CURRENT_PATCHLEVEL"]))
    update_kernel_info("CURRENT_SUBLEVEL:" + str(version_info["CURRENT_SUBLEVEL"]))
    update_kernel_info("LAUNCHVERSION:" + str(version_info["LAUNCHVERSION"]))
    update_kernel_info("KMI_GENERATION:" + str(version_info["KMI_GENERATION"]))
    update_kernel_info("COMMIT:" + str(version_info["COMMIT"]))


def update_kernel_info(new_info):
    global kernel_info
    kernel_info.append(new_info)


def write_info_to_file(path):
    global kernel_info
    dir_name = os.path.dirname(path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)
    with open(path, 'w') as f:
        f.write("\n".join(kernel_info))


def calculate_file_md5(file_path):
    """
    Calculate the MD5 value of a single file.

    :param file_path: Path to the file
    :return: MD5 value of the file as a string. Returns None if the file does not exist or
    an error occurs while reading.
    """
    if not os.path.isfile(file_path):
        return None

    try:
        md5 = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                md5.update(chunk)
        return md5.hexdigest()
    except Exception as e:
        print("Error reading file {}: {}".format(file_path, e))
        return None


def write_md5_to_file(md5_results, output_file):
    """
    Write the MD5 results to a specified file and return the MD5 value of the output file.

    :param md5_results: Dictionary of MD5 results
    :param output_file: Path to the output file
    :return: MD5 value of the output file
    """
    with open(output_file, "a") as out_f:
        for file_name, md5sum in md5_results.items():
            out_f.write("{}: {}\n".format(file_name, md5sum))

    # Calculate the MD5 value of the output file itself
    return calculate_file_md5(output_file)


def calculate_files_md5sum(file_paths, output_file=None, base_path="."):
    """
    Calculate the MD5sum values of multiple files and output the results to a specified file (optional), or return the
    MD5sum of a single file.

    :param file_paths: List of file names
    :param output_file: Path to the output file (optional)
    :param base_path: Path to the directory containing the files (default is the current directory)
    :return: If a single file, return the MD5 value of the file. If multiple files and an output file is specified,
     return the MD5 value of the output file.
    """

    md5_results = {}

    for file_name in file_paths:
        file_path = os.path.join(base_path, file_name)
        md5sum = calculate_file_md5(file_path)
        md5_results[file_name] = md5sum if md5sum is not None else "ERROR"

    if output_file:
        return write_md5_to_file(md5_results, output_file)

    if len(file_paths) == 1:
        # If only one file, return the MD5 value of that file
        return list(md5_results.values())[0]

    return md5_results


def delete_files_and_dirs(path, delete_list):
    # Ensure the provided directory exists
    if not os.path.exists(path):
        print('The provided directory does not exist.')
        return

    # Iterate over the list of files/directories to delete
    for item in delete_list:
        item_path = os.path.join(path, item)
        # If the item exists, delete it
        if os.path.exists(item_path):
            # If it's a directory, use shutil.rmtree() to delete it
            if os.path.isdir(item_path):
                shutil.rmtree(item_path)
            # If it's a file, use os.remove() to delete it
            else:
                os.remove(item_path)


def calculate_sha256(input_string):
    """
    Calculate and print the SHA256 hash value of a string.

    Parameters:
    input_string (str): The string to calculate the hash value for

    Returns:
    str: The SHA256 hash value of the string
    """
    # Create a SHA256 hash object
    sha256_hash = hashlib.sha256()

    # Update the hash object with the byte representation of the string
    sha256_hash.update(input_string.encode('utf-8'))

    # Get the hexadecimal hash value
    sha256_digest = sha256_hash.hexdigest()

    # Print the hash value
    print("calculate sha256: " + sha256_digest)

    # Return the hash value
    return sha256_digest


def find_id_in_xml(xml_filename, search_id):
    """
    Read the content of an XML file and search for elements with an id attribute matching the provided string.

    :param xml_filename: The name of the XML file
    :param search_id: The string to search for
    :return: Returns (id, bug) if found, otherwise returns False
    """
    try:
        tree = ET.parse(xml_filename)
        root = tree.getroot()

        # Find all <build> elements
        for build in root.iter('build'):
            # Check if the id attribute of the current <build> element matches the search string
            if build.get('id') == search_id:
                bug = build.get('bug')
                print("Found id: {}, bug: {}".format(search_id, bug))
                return search_id, bug

        # If no matching id is found
        return False

    except ET.ParseError:
        print("Error: Failed to parse the XML file.")
        return False
    except FileNotFoundError:
        print("Error: The file was not found.")
        return False


def check_image_type(input_string):
    if '-abogki' in input_string:
        return 'ogki'
    elif '-ab' in input_string and '-abogki' not in input_string:
        return 'gki'
    elif '-o-' in input_string:
        return 'oki'
    else:
        return 'unknown'


def ogki_xts_approve_check(args, version):
    update_info_parts = []
    approved = "fail"
    google_bug_id = "0"

    sha256_result = calculate_sha256(version)
    image_type = check_image_type(version)

    if os.path.exists(args.approved):

        result = find_id_in_xml(args.approved, sha256_result)
        if result:
            approved = "pass"
            google_bug_id = result[1]
        else:
            if image_type == "ogki":
                print("No matching id found.")
                approved = "fail"
                google_bug_id = "0"
            else:
                approved = "ignore"
                google_bug_id = "ignore"

    update_info_parts.append("LINUX_KERNEL_FULL_VERSION_SHA256:" + sha256_result)
    update_info_parts.append("LINUX_KERNEL_FULL_VERSION_OGKI_VTS_CHECK:" + approved)
    update_info_parts.append("google_bug_id:" + google_bug_id)
    update_info_parts.append("image_type:" + image_type)

    message = "\n".join(update_info_parts) + "\n"
    print(message)

    update_kernel_info(message)

    approve_data = [
        version,
        image_type
    ]
    return approve_data


def find_kernel_lifetime_info(xml_filename, branch_name, release_version):
    """
    Read the content of an XML file and search for the specified branch and release version information.

    :param xml_filename: The name of the XML file
    :param branch_name: The branch name to search for
    :param release_version: The release version to search for
    :return: If found, return (branch_info, release_info), otherwise return False
    """
    try:
        tree = ET.parse(xml_filename)
        root = tree.getroot()

        branch_info = release_info = None

        # Find all <branch> elements
        for branch in root.iter('branch'):
            # Check if the name attribute of the current <branch> element matches the branch_name
            if branch.get('name') == branch_name:
                branch_info = {
                    'name': branch.get('name'),
                    'version': branch.get('version'),
                    'launch': branch.get('launch'),
                    'eol': branch.get('eol')
                }
                print("\nFound branch: name: {name}, version: {version}, launch: {launch}, eol: {eol}".format(
                    name=branch_info.get('name'),
                    version=branch_info.get('version'),
                    launch=branch_info.get('launch'),
                    eol=branch_info.get('eol')
                ))
                # Find all <release> elements
                for release in branch.findall('.//release'):
                    # Check if the version attribute of the current <release> element matches the release_version
                    if release.get('version') == release_version:
                        release_info = {
                            'version': release.get('version', '0'),
                            'launch': release.get('launch', '0'),
                            'eol': release.get('eol', '0')
                        }
                        print("Found release: version: {version}, launch: {launch}, eol: {eol}".format(
                            version=release_info.get('version', '0'),
                            launch=release_info.get('launch', '0'),
                            eol=release_info.get('eol', '0')
                        ))

                if release_info:
                    return branch_info, release_info
                elif branch_info:
                    return branch_info, {'version': '0', 'launch': '0', 'eol': '0'}
        # If no matching branch or release is found
        return False

    except ET.ParseError:
        print("Error: Failed to parse the XML file.")
        return False
    except FileNotFoundError:
        print("Error: The file was not found.")
        return False


def format_as_text(value):
    """
    Formats the value as text to avoid automatic format conversion in Excel.
    :param value: The value to be formatted.
    :return: The formatted text.
    """

    def is_date_string(s):
        """
        Check if the string is a date in the format 'YYYY-MM-DD'.
        """
        try:
            datetime.strptime(s, '%Y-%m-%d')
            return True
        except ValueError:
            return False

    def is_scientific_notation(s):
        """
        Check if the string is a valid scientific notation.
        """
        try:
            float(s)
            # Check if the string contains 'E' or 'e' and is a valid scientific notation
            if 'E' in s.upper():
                return True
            return False
        except ValueError:
            return False

    def needs_preservation(s):
        """
        Check if the string is a number with precision that should be preserved
        or if it's a date that should not be interpreted by Excel.
        """
        if is_date_string(s):
            return True

        if is_scientific_notation(s):
            return True

        try:
            # Try converting to float to see if it's a number
            float_val = float(s)
            # If it contains a decimal point and a non-zero digit after it, preserve it
            if '.' in s and s[-1] == '0':
                return True
            return False
        except ValueError:
            return False

    if isinstance(value, str):
        # Check if the string needs preservation
        if needs_preservation(value):
            return "'{}'".format(value)  # Add single quote to preserve date/numeric precision
        # Return the original string for all other cases
        return value
    elif isinstance(value, (int, float)):
        # Return numeric values as they are
        return str(value)
    # For other types, convert to string
    return str(value)


def modify_filename(file_name, suffix):
    """
    Args:
        suffix:
        file_name:
    """
    # Use pathlib.Path to get the file extension
    file_path = Path(file_name)
    file_suffix = file_path.suffix

    # Check if the file extension is .xlsx
    if file_suffix.lower() != suffix:
        # If not .xlsx, modify the file name
        new_file_name = file_path.with_suffix(suffix)
        print(f"File name modified: {new_file_name}")
    else:
        # If already .xlsx, keep the original name
        new_file_name = file_name
        print(f"File name is already {suffix}: {new_file_name}")
    return str(new_file_name)


def write_build_info_to_file(filename, headers, data, update_mode=True, key_columns=None):
    """
    Write data to a file (Excel or CSV) with the given headers.

    :param filename: The name of the file to write to.
    :param headers: A list of column headers.
    :param data: A list of values to write in one row.
    :param update_mode: Boolean to specify if update mode is enabled.
    :param key_columns: List of key column names to check for updates.
    """

    if key_columns is None:
        key_columns = []

    data = [format_as_text(value) for value in data]
    print("data: {}".format(data))
    if filename.endswith('.xlsx'):
        # Check if the Excel file exists
        file_exists = os.path.exists(filename)
        """
        if file_exists:
            # Load the existing workbook
            wb = load_workbook(filename)
            ws = wb.active

            # Read the header row
            headers_in_file = [cell.value for cell in ws[1]]
            if headers_in_file != headers:
                print("Provided headers do not match with existing Excel headers. Updating headers to: {}".format(headers))
                # Update headers and write all records with new headers
                ws.delete_rows(1)
                ws.append(headers)
                # Re-read the file with new headers
                headers_in_file = headers

            if update_mode:
                # Check if there is a completely identical row
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if list(row) == data:
                        # Completely identical row found, no need to update
                        wb.save(filename)
                        return

                # Check if there is a row that matches the key columns
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if all(row[headers_in_file.index(col)] == data[headers.index(col)] for col in key_columns):
                        # Update this row
                        for col_idx, value in enumerate(data):
                            ws.cell(row=row_idx, column=col_idx + 1, value=value)
                        wb.save(filename)
                        return

        else:
            # Create a new Excel file and write the header
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        # Append the data as a new row
        ws.append(data)
        wb.save(filename)
        """
    else:
        # For CSV format
        file_exists = os.path.isfile(filename)
        records = []
        headers_in_file = []
        if file_exists:
            # Read the existing records
            with open(filename, mode='r', newline='') as file:
                reader = csv.reader(file)
                headers_in_file = next(reader)  # Read the header row
                records = list(reader)

            if headers_in_file != headers:
                print(f"Provided headers do not match with existing CSV headers. Updating headers to: {headers}")
                # Update headers and write all records with new headers
                with open(filename, mode='w', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow(headers)
                    writer.writerows(records)
                # Re-read the file with new headers
                with open(filename, mode='r', newline='') as file:
                    reader = csv.reader(file)
                    headers_in_file = next(reader)  # Read the updated header row
                    records = list(reader)

            if update_mode:
                # Check if there is a completely identical row
                if data in records:
                    # Completely identical row found, no need to update
                    return

                # Check if there is a row that matches the key columns
                for i, row in enumerate(records):
                    if all(row[headers.index(col)] == data[headers.index(col)] for col in key_columns):
                        # Update this row
                        records[i] = data
                        with open(filename, mode='w', newline='') as file:
                            writer = csv.writer(file)
                            writer.writerow(headers)
                            writer.writerows(records)
                        return

        # Append the data as a new row in the CSV
        with open(filename, mode='a', newline='') as file:
            writer = csv.writer(file)
            if not file_exists:
                writer.writerow(headers)
            writer.writerow(data)


def days_until_eol(eol_date_str, current_date_str, default_value=0):
    """
    Calculate the number of days remaining until the end of life (EOL) date.

    Parameters:
    eol_date_str (str): The end of life date in the format "YYYY-MM-DD".
    current_date_str (str): The current date in the format "YYYY-MM-DD".
    default_value (any): The default value to return in case of invalid date format or calculation error (default is -1)

    Returns:
    int: The number of days remaining until the EOL date, or default_value if the date is invalid or calculation fails.
    """
    date_format = "%Y-%m-%d"

    try:
        eol_date = datetime.strptime(eol_date_str, date_format)
        current_date = datetime.strptime(current_date_str, date_format)

        # Calculate the number of days remaining
        delta = eol_date - current_date
        return delta.days

    except ValueError:
        return default_value
    except Exception as e:  # Catch any other exceptions
        print("An error occurred: {}".format(e))
        return default_value


def kernel_lifetime_check(args, branch_name, release_version, build_date):
    if branch_name is None or release_version is None or not os.path.exists(args.lifetimes):
        return ['0', '0', '0', '0', '0', '0', '0', '0', '0'], '0', '0'
    result = find_kernel_lifetime_info(args.lifetimes, branch_name, release_version)
    if result:
        branch_info, release_info = result
        print("Branch Info: {name}, {version}, {launch}, {eol}".format(
            name=branch_info.get('name', '0'),
            version=branch_info.get('version', '0'),
            launch=branch_info.get('launch', '0'),
            eol=branch_info.get('eol', '0')
        ))
        print("Release Info: {version}, {launch}, {eol}".format(
            version=release_info.get('version', '0'),
            launch=release_info.get('launch', '0'),
            eol=release_info.get('eol', '0')
        ))
    else:
        print("No matching branch or release found.")
        branch_info, release_info = {}, {}

    eol = release_info.get('eol', '0')
    print("build_date {} {}".format(eol, build_date))
    days_left = days_until_eol(eol, build_date)
    print("days_left {}".format(days_left))
    print("eol {}".format(eol))
    if eol == '0':
        ack_lts_message = "default"
    elif days_left > 0:
        ack_lts_message = "pass"
    else:
        ack_lts_message = "error"

    update_info = "ACK_LTS_SPL_CHECK:" + ack_lts_message
    print(update_info + "\n")
    update_kernel_info(update_info)

    lifetime_data = [
        release_info.get('version', '0'),
        release_info.get('launch', '0'),
        eol,
        build_date,
        days_left,
        branch_info.get('name', '0'),
        branch_info.get('version', '0'),
        branch_info.get('launch', '0'),
        branch_info.get('eol', '0')
    ]
    return lifetime_data, days_left, eol


def find_current_linux_version(args):
    print('\ndownload image')
    handle_boot_img_path(args.bootimg, args.out)

    if not os.path.exists(args.out_bootimg):
        print("{} File does not exist.ignore".format(args.out_bootimg))
        return None
    version = common.get_boot_img_version(args.unpack_bootimg, args.out_bootimg, args.out_unpack,
                                          args.out_kernel, args.lz4)
    update_info = "LINUX_KERNEL_FULL_VERSION:" + version + "\n"
    print(update_info)
    update_kernel_info(update_info)
    return version


def clone_single_branch(repo_url, branch, destination):
    """
    Clone a specified branch of a Git repository and only fetch the latest commit.

    :param repo_url: URL of the repository
    :param branch: Name of the branch to clone
    :param destination: Name of the local directory to clone into
    """
    try:
        # Check if the destination directory exists, and delete it if it does
        if os.path.exists(destination):
            print("{} already exists".format(destination))
            subprocess.run(["rm", "-rf", destination], check=True)

        # Use shallow clone to clone a single branch and only include the latest commit
        subprocess.run([
            "git", "clone", "--depth", "1", "--branch", branch, repo_url, destination
        ], check=True)

        print("gerrit {} branch {} cloned to {}".format(repo_url, branch, destination))
    except subprocess.CalledProcessError as e:
        print("clone error: {}".format(e))


def clone_kernel_config(args):
    approved = os.path.normpath(os.path.join(args.ack_config, "approved-ogki-builds.xml"))
    lifetimes = os.path.normpath(os.path.join(args.ack_config, "kernel-lifetimes.xml"))
    approved_target = os.path.normpath(args.approved)
    lifetimes_target = os.path.normpath(args.lifetimes)

    if not os.path.exists(approved) or not os.path.exists(lifetimes):
        repo_url = "ssh://gerrit_url:29418/kernel/configs"
        branch = "upstream/google/main"
        clone_single_branch(repo_url, branch, args.ack_config)

    if approved != approved_target:
        if os.path.exists(approved):
            shutil.copy2(approved, approved_target)
            print(f"File copied successfully from {approved} to {approved_target}")
        else:
            print(f"Source file does not exist: {approved}")

    if lifetimes != lifetimes_target:
        if os.path.exists(lifetimes):
            shutil.copy2(lifetimes, lifetimes_target)
            print(f"File copied successfully from {lifetimes} to {lifetimes_target}")
        else:
            print(f"Source file does not exist: {lifetimes}")


def extract_fixed_position_value_from_strings(strings, position):
    """
    Extracts the value at a fixed position from the given URL.

    :param strings: The strings from which to extract the value.
    :param position: The position of the value to extract (0-based index).
    :return: The value at the specified position or None if not found.
    """
    try:
        # Split the URL by '/'
        parts = strings.split('/')

        # Check if the position is within the bounds of the list
        if position < len(parts):
            return parts[position]
        else:
            return '0'
    except Exception as e:
        # Handle any unexpected errors
        print(f"An error occurred: {e}")
        return '0'


def list_to_dict(build_json_data, keys):
    """
    Converts a list to a dictionary using the provided keys.

    :param build_json_data: The list of values to convert.
    :param keys: The list of keys to use for the dictionary.
    :return: A dictionary with keys mapped to the corresponding values from the list.
    """
    if len(build_json_data) != len(keys):
        raise ValueError("The length of build_json_data and keys must be the same.")

    return dict(zip(keys, build_json_data))


def build_json_check_and_write_data(json_file):
    """
    Extract data from a specified JSON file and write it to CSV and Excel files.

    :param json_file: Path to the input JSON file
    """

    if os.path.exists(json_file):

        with open(json_file, 'r') as f:
            data = json.load(f)

        project = data.get("ofp", {}).get("build_cfg", {}).get("PRODUCT_VARIANT", "0")
        branch = data.get("vnd", {}).get("OPLUS_CCM_MANIFEST_BRANCH", "0")
        area = data.get("ofp", {}).get("build_cfg", {}).get("OPLUS_AREA", "0")
        prebuilt = data.get("vnd", {}).get("build_cfg", {}).get("OPLUS_USE_PREBUILT_BOOTIMAGE", "0")
        version_type = extract_fixed_position_value_from_strings(data.get("ofp", {}).get("super_map", "0"), 5)
        # print(f"version_type: {version_type}")
        # Extract data items with validity checks and default values
        build_json_data = [
            data.get("vnd", {}).get("OPLUS_CCM_SOC", "0"),
            data.get("vnd", {}).get("OPLUS_CCM_PLATFORM", "0"),
            data.get("ofp", {}).get("build_cfg", {}).get("OPLUS_BRAND", "0"),
            project,
            branch,
            area,
            prebuilt,
            version_type
        ]

    else:
        build_json_data = [
            "0",
            "0",
            "0",
            "0",
            "0",
            "0",
            "0",
            "0"
        ]

    keys = [
        'vendor',
        'platform',
        'brand',
        'project',
        'branch',
        'area',
        'prebuilt',
        'version_type'
    ]

    # 转换为字典
    build_json_dict = list_to_dict(build_json_data, keys)
    # print(build_json_dict)
    return build_json_data, build_json_dict


def get_kernel_upstream(file_path):
    if not os.path.exists(file_path):
        return None, None

    tree = ET.parse(file_path)
    root = tree.getroot()

    # 定义匹配的 name 模式列表，支持通配符
    match_names = ['kernel/kernel-*', 'quark/kernel-*', 'kernel/common']

    for pattern in match_names:
        for project in root.findall('.//project'):
            name = project.get('name')
            if fnmatch.fnmatch(name, pattern):
                ori_upstream = project.get('ori_upstream')
                upstream = project.get('upstream')
                if ori_upstream:
                    return name, ori_upstream
                elif upstream:
                    return name, upstream

    return None, None


def get_jfrog_url_info(args):
    file_path = os.path.join(args.out, "merge_component_manifests.xml")
    name, upstream_value = get_kernel_upstream(file_path)
    if name and upstream_value:
        print(f"Name: {name}, Upstream: {upstream_value}")
    else:
        print("No matching project found.")

    bootimg = args.bootimg if args.bootimg else '0'

    jfrog_url_info = [
        name if name else '0',
        upstream_value if upstream_value else '0',
        bootimg
    ]
    return jfrog_url_info, upstream_value


def check_ack_conditions(args, version, version_info):
    # 检查 version 和 version_info 是否为空字符串
    if not version or not version.strip():
        print("warning: version is an empty string.")
        return False

    # 检查文件 args.build_json 是否存在
    if not os.path.exists(args.build_json):
        print(f"warning: File {args.build_json} does not exist.")
        return False

    return True


def get_project_info_from_sever(args, project_number,gerrit_branch):
    data_dict = get_prj_user_number_cmd(args, project_number)
    create_user_no = data_dict.get('CREATE_USER_NO', '0')
    pm_list = data_dict.get('SOFTWARE_PM', '0')
    cmo_list = data_dict.get('CMO', '0')
    project_name = data_dict.get('PROJECT_NAME', '0')

    # 将列表转换为 | 分隔的字符串，并提供默认值
    create_user_no_str = '|'.join(create_user_no) if create_user_no else '0'
    software_pm_str = '|'.join(pm_list) if pm_list else '0'
    cmo_str = '|'.join(cmo_list) if cmo_list else '0'

    # 调用函数并获取字典
    project_info = read_project_info_config(args, project_number, gerrit_branch)
    config_userid = project_info.get('config_userid', '0')
    ignore_date_start = project_info.get('ignore_date_start', '0')
    ignore_date_end = project_info.get('ignore_date_end', '0')
    warning_delta = int(project_info.get('warning_delta', '0'))
    error_delta = int(project_info.get('error_delta', '0'))
    warning_duty = int(project_info.get('warning_duty', '0'))
    error_duty = int(project_info.get('error_duty', '0'))
    enable_error = project_info.get('enable_error', '0')

    # 将字典中的值转换为适当的类型
    project_info_dict = {
        'config_userid': config_userid,
        'ignore_date_start': ignore_date_start,
        'ignore_date_end': ignore_date_end,
        'warning_delta': warning_delta,
        'error_delta': error_delta,
        'warning_duty': warning_duty,
        'error_duty': error_duty,
        'enable_error': enable_error
    }
    # 返回 prj_data
    project_info_array = [
        project_name,
        config_userid,
        software_pm_str,
        cmo_str,
        create_user_no_str,
        ignore_date_start,
        ignore_date_end,
        warning_delta,
        error_delta,
        warning_duty,
        error_duty,
        enable_error
    ]

    return project_info_array, project_info_dict


def generate_ack_data_list_save(args, version, version_info, build_date):
    clone_kernel_config(args)

    if not check_ack_conditions(args, version, version_info):
        # print("One or more conditions are not met. Ignoring.")
        return None

    required_keys = ["CURRENT_KERNEL_VERSION", "LAUNCHVERSION", "CURRENT_VERSION", "CURRENT_PATCHLEVEL"]
    if all(key in version_info for key in required_keys):
        release_version = version_info["CURRENT_KERNEL_VERSION"]
        branch_name = "android{}-{}.{}".format(version_info["LAUNCHVERSION"],
                                               version_info["CURRENT_VERSION"],
                                               version_info["CURRENT_PATCHLEVEL"])
    else:
        release_version = '0'
        branch_name = '0'

    build_json_data, build_json_dict = build_json_check_and_write_data(args.build_json)
    project = build_json_dict['project']
    area = build_json_dict['area']
    branch = build_json_dict['branch']
    prebuilt = build_json_dict['prebuilt']
    version_type = build_json_dict['version_type']
    print("build_json_data " + str(build_json_data))

    jfrog_url_info, gerrit_branch = get_jfrog_url_info(args)
    print("jfrog_url_info " + str(jfrog_url_info))

    project_info_data, project_info_dict = get_project_info_from_sever(args, project, gerrit_branch)
    print("\nprj_info_date " + str(project_info_data))

    lifetime_data, days_left, release_eol = kernel_lifetime_check(args, branch_name, release_version, build_date)
    print("lifetime_data " + str(lifetime_data))

    approve_data = ogki_xts_approve_check(args, version)
    print("approve_data " + str(approve_data))

    final_data = build_json_data + jfrog_url_info + lifetime_data + approve_data + project_info_data

    headers = ['vendor', 'platform', 'brand', 'project', 'ccm_cfg_branch', 'area',
               'prebuilt', 'version_type',
               'gerrit_projects', 'gerrit_branch', 'jfrog_url',
               'release_version', 'release_launch',
               'release_eol', 'build_date', 'days_eol',
               'main_branch', 'main_version', 'main_launch', 'main_eol',
               'kernel_version_string', 'image_type',
               'name', 'config_userid', 'pm_id', 'cmo_id', 'grs_userid', 'ignore_date_start',
               'ignore_date_end', 'warning_delta', 'error_delta', 'warning_duty', 'error_duty', 'enable_error'
               ]

    key_columns = ['vendor', 'platform', 'brand', 'project', 'ccm_cfg_branch', 'area']

    error_branch = check_branch(branch)
    if not error_branch:
        print(f"ignore branch {branch}")
        return None

    days_range = days_until_eol(str(args.current_date), str(build_date))
    if release_eol !='0' and abs(days_range) < 7:
        write_build_info_to_file(args.ack_all_info, headers, final_data, update_mode=True, key_columns=key_columns)
        #write_build_info_to_file(modify_filename(args.ack_all_info, '.xlsx'),headers,
        #final_data,update_mode=True, key_columns=key_columns)
    else:
        print(f"release_eol {release_eol} current_date {args.current_date} build_date {build_date} days_range {days_range} ignore")

    ignore_date_start = project_info_dict['ignore_date_start']
    ignore_date_end = project_info_dict['ignore_date_end']
    error_delta = project_info_dict['error_delta']
    enable_error = project_info_dict['enable_error']

    # 将字典中的值转换为适当的类型
    build_info_dict = {
        'project': project,
        'days_left': days_left,
        'build_date': build_date,
        'release_eol': release_eol,
        'ignore_date_start': ignore_date_start,
        'ignore_date_end': ignore_date_end,
        'error_delta': error_delta,
        'branch': branch,
        'area': area,
        'prebuilt': prebuilt,
        'version_type': version_type,
        'enable_error': enable_error
    }
    # print(f"build_info_dict {build_info_dict}")
    return build_info_dict


def check_branch(branch):
    # pattern = r'\b(master|release)\b'
    pattern = r'\b(dev|master|release)\b'
    error_branch = bool(re.search(pattern, branch, re.IGNORECASE))
    return error_branch


def check_version_type(version_type):
    """
    Checks if the given version_type is 'dailybuild' (case-insensitive).

    :param version_type: The version type to check.
    :return: True if version_type is 'dailybuild', otherwise False.
    """
    if version_type is None:
        return False

    return version_type.lower() == 'dailybuild'


def check_ogki_prj_approve(ogki_prj_approve, project):
    if not os.path.exists(ogki_prj_approve):
        print(f"File {ogki_prj_approve} does not exist.")
        return False

    with open(ogki_prj_approve, mode='r', newline='') as file:
        reader = csv.DictReader(file)
        for row in reader:
            if row['project'] == project:
                print("{} in ogki ignore list, auto approve".format(project))
                return True
    return False

def ogki_error_check(args, project, branch, area, prebuilt):
    """
    Checks if the given branch, area, and prebuilt meet specific conditions.

    :param branch: The branch name.
    :param area: The area name.
    :param prebuilt: The prebuilt name.
    :return: False if all conditions are met, otherwise True.
    """
    # if check_ogki_prj_approve(args.ogki_prj_approve, project):
    #    return False

    # Check if branch contains 'release'
    if 'release' in branch.lower():
        # Check if area export or gdpr
        if 'export' in area.lower() or 'gdpr' in area.lower():
            # Check if prebuilt contains 'ogki'
            if 'ogki' in prebuilt.lower():
                return True

    return False


def check_build_status(args, build_info_dict):
    if not build_info_dict:
        print("warning: build_info_dict is empty or None.")
        return

    # 从 build_info_dict 中获取必要的变量
    project = build_info_dict.get('project', 0)
    days_left = build_info_dict.get('days_left', 0)
    build_date = build_info_dict.get('build_date', '0')
    release_eol = build_info_dict.get('release_eol', '0')
    ignore_date_start = build_info_dict.get('ignore_date_start', args.ignore_date_start)
    ignore_date_end = build_info_dict.get('ignore_date_end', args.ignore_date_end)
    error_delta = build_info_dict.get('error_delta', args.error_delta)
    enable_error = build_info_dict.get('enable_error', args.enable_error)
    branch = build_info_dict.get('branch', '0')
    area = build_info_dict.get('area', '0')
    prebuilt = build_info_dict.get('prebuilt', '0')

    # 将字符串转换为适当的类型
    try:
        days_left = int(days_left)
    except (ValueError, TypeError):
        print("Error: Failed to convert days_left to integer. Using default value 0.")
        days_left = 0

    error_delta = int(error_delta) if error_delta is not None else int(args.error_delta)
    enable_error = int(enable_error) if enable_error is not None else int(args.enable_error)

    # 清洗并转换日期字符串
    is_valid, build_date = is_valid_date(build_date)
    is_valid, release_eol = is_valid_date(release_eol)
    is_valid, ignore_date_start = is_valid_date(ignore_date_start)
    is_valid, ignore_date_end = is_valid_date(ignore_date_end)

    # 打印所有相关的变量
    """
    print(f"\ndays_left: {days_left}")
    print(f"build_date: {build_date}")
    print(f"release_eol: {release_eol}")
    print(f"ignore_date_start: {ignore_date_start}")
    print(f"ignore_date_end: {ignore_date_end}")
    print(f"error_delta: {error_delta}")
    print(f"branch: {branch}\n")
    """
    is_within_ignore_range, ignore_days_left = is_date_within_range_and_days_left(str(build_date),
                                                                                  str(ignore_date_start),
                                                                                  str(ignore_date_end))
    """
    print(f"current date:[{build_date}] is within the date range [{ignore_date_start}] to [{ignore_date_end}] "
          f"is_within_ignore_range[{is_within_ignore_range}]")
    print(f"ignore_days_left[{ignore_days_left}] until[{ignore_date_end}] \n")
    """
    need_error = (1 < days_left < error_delta)
    error_branch = check_branch(branch)
    daily_build = check_version_type(build_info_dict.get('version_type', '0'))
    error_ogki = ogki_error_check(args, project, branch, area, prebuilt)
    """
    print(f"need_error:{need_error}\nis_within_ignore_range:{is_within_ignore_range}\n"
          f"enable_error: {enable_error} \nerror_branch: {error_branch}\nerror_ogki: {error_ogki}\n"
          f"daily_build: {daily_build}")
    """
    # 检查条件
    if (need_error and daily_build) or error_ogki:
        if (not is_within_ignore_range and error_branch and enable_error) or error_ogki:
            print(
                f"Error: Days left ({days_left}) is less than error delta ({error_delta})"
                f" and current date ({build_date}) is not in the ignore date range "
                f"({ignore_date_start} to {ignore_date_end}) and enable_error is {enable_error}"
                f" error_ogki is {error_ogki}")
            print("Build status check failed. Exiting.")
            # sys.exit(1)
        else:
            print(f"need_error {need_error} but in ignore status")
    else:
        print("check default pass")


def extract_and_split_valid(input_str):
    """
    从输入字符串中提取并分割有效的用户ID。

    :param input_str: 输入字符串
    :return: 分割后的用户ID列表
    """
    if not input_str or input_str == '0':
        return []
    parts = re.split(r'[|,\s]+', input_str)
    return [part.strip() for part in parts if part.strip()]


def standardize_delimiters(input_str):
    """
    将输入字符串中的所有分隔符（空格、|、逗号）标准化为单个 |，
    并去除多余的 | 和字符串两端的 |。

    :param input_str: 输入字符串
    :return: 标准化后的字符串
    """
    # 使用正则表达式将所有分隔符（空格、|、逗号）替换为单个 |
    result = re.sub(r'[|,\s]+', '|', input_str)
    # 去除多余的 |，确保没有连续的 ||
    result = re.sub(r'\|+', '|', result)
    # 去除字符串两端的 |
    result = result.strip('|')
    return result


def save_config_project_info(args, gerrit_branch, config_userid):

    warning_delta = args.warning_delta
    error_delta = args.error_delta
    warning_duty = args.warning_duty
    error_duty = args.error_duty
    enable_error = args.enable_error

    start_valid, ignore_date_start = is_valid_date(args.ignore_date_start)
    end_valid, ignore_date_end = is_valid_date(args.ignore_date_end)
    # print(f"\nstart_valid: {start_valid}, ignore_date_start: {ignore_date_start}")
    # print(f"end_valid: {end_valid}, ignore_date_end: {ignore_date_end}")
    if not start_valid or not end_valid:
        ignore_date_start = '0'
        ignore_date_end = '0'

    days_left = days_until_eol(str(ignore_date_end), str(ignore_date_start))
    # print("days_left {}".format(days_left))
    if days_left > int(args.error_delta) / 2:
        ignore_date_end = adjust_date_by_days(str(ignore_date_start), int(args.error_delta) / 2, 'add')
        print("local_date {}".format(ignore_date_end))

    # print(f"ignore_date_start {ignore_date_start}: ignore_date_end {ignore_date_end}")

    # 定义CSV文件的列名
    headers = ['gerrit_branch', 'config_userid', 'ignore_date_start', 'ignore_date_end', 'warning_delta',
               'error_delta', 'warning_duty', 'error_duty', 'enable_error']
    file_exists = os.path.exists(args.prj_cfg_info)

    updated = False
    rows = []

    if file_exists:
        # 读入现有的CSV文件
        with open(args.prj_cfg_info, mode='r', newline='') as file:
            reader = csv.DictReader(file)
            rows = list(reader)

        # 查找是否存在相同的项目并更新
        for row in rows:
            if row.get('gerrit_branch', '0') == gerrit_branch:
                # print("\nBefore Update:", row)
                if config_userid:
                    row['config_userid'] = config_userid
                if ignore_date_start:
                    row['ignore_date_start'] = ignore_date_start
                if ignore_date_end:
                    row['ignore_date_end'] = ignore_date_end
                if warning_delta:
                    row['warning_delta'] = warning_delta
                if error_delta:
                    row['error_delta'] = error_delta
                if warning_duty:
                    row['warning_duty'] = warning_duty
                if error_duty:
                    row['error_duty'] = error_duty
                if enable_error:
                    row['enable_error'] = enable_error
                # print("After Update:", row)
                updated = True
                break

    if not updated:
        # 如果不存在相同的项目，则添加新的条目
        new_row = {
            'gerrit_branch': gerrit_branch,
            'config_userid': config_userid,
            'ignore_date_start': ignore_date_start,
            'ignore_date_end': ignore_date_end,
            'warning_delta': warning_delta,
            'error_delta': error_delta,
            'warning_duty': warning_duty,
            'error_duty': error_duty,
            'enable_error': enable_error
        }
        rows.append(new_row)
        # print("New Entry:", new_row)

    # print("info:\n", rows)
    # 写入更新或新创建的内容到CSV文件
    with open(args.prj_cfg_info, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            try:
                # 过滤掉不在headers中的字段
                row_filtered = {key: format_as_text(row[key]) for key in headers if key in row}
                writer.writerow(row_filtered)
            except ValueError as e:
                print(f"Error writing row {row}: {e}")
                continue


def read_project_info_config(args, project, gerrit_branch):
    """
    读取 CSV 文件，查找与传递的 project 相同的行，并返回各列的内容。
    """
    file_path = args.prj_cfg_info

    # 定义默认值
    default_values = {
        'config_userid': args.config_userid,
        'ignore_date_start': args.ignore_date_start,
        'ignore_date_end': args.ignore_date_end,
        'warning_delta': args.warning_delta,
        'error_delta': args.error_delta,
        'warning_duty': args.warning_duty,
        'error_duty': args.error_duty,
        'enable_error': args.enable_error
    }

    if not os.path.exists(file_path):
        print(f"file  {file_path} not exist")
        return default_values

    # 读取 CSV 文件
    with open(file_path, mode='r', newline='') as file:
        reader = csv.DictReader(file)
        rows = list(reader)

    # 查找是否存在相同的项目
    for row in rows:
        if row.get('gerrit_branch', '0') == gerrit_branch:
            return {
                'config_userid': row.get('config_userid', args.config_userid),
                'ignore_date_start': row.get('ignore_date_start', args.ignore_date_start),
                'ignore_date_end': row.get('ignore_date_end', args.ignore_date_end),
                'warning_delta': row.get('warning_delta', args.warning_delta),
                'error_delta': row.get('error_delta', args.error_delta),
                'warning_duty': row.get('warning_duty', args.warning_duty),
                'error_duty': row.get('error_duty', args.error_duty),
                'enable_error': row.get('enable_error', args.enable_error)
            }

    print(f"not find project {project} config info in {file_path} use default")
    return default_values


def adjust_date_by_days(date_string, days, operation='subtract'):
    """
    从给定的日期字符串中增加或减去指定的天数。

    :param date_string: 日期字符串，格式为 'YYYY-MM-DD'
    :param days: 要增加或减去的天数
    :param operation: 操作类型，'add' 表示增加，'subtract' 表示减少，默认为 'subtract'
    :return: 调整后的日期字符串，格式为 'YYYY-MM-DD'，如果输入无效则返回 '0'
    """
    try:
        # 将日期字符串转换为 datetime 对象
        date_object = datetime.strptime(date_string, '%Y-%m-%d')
    except ValueError:
        return '0'

    # 根据操作类型调整日期
    if operation == 'add':
        new_date_object = date_object + timedelta(days=days)
    else:
        new_date_object = date_object - timedelta(days=days)

    # 将新的 datetime 对象转换回字符串
    new_date_string = new_date_object.strftime('%Y-%m-%d')

    return new_date_string


def is_valid_date(date_str):
    """
    判断输入的字符串是否是有效的日期格式，并返回相应的结果。
    如果有效，返回标准的 YYYY-MM-DD 格式；如果无效，返回 '0'。

    :param date_str: 输入的日期字符串
    :return: 标准的 YYYY-MM-DD 格式字符串或 '0'
    """
    date_str = date_str.strip().strip("'").strip('"')
    # 定义可能的日期格式
    date_formats = [
        "%Y-%m-%d",  # YYYY-MM-DD
        "%Y/%m/%d",  # YYYY/MM/DD
        "%Y.%m.%d",  # YYYY.MM.DD
        "%m/%d/%Y",  # MM/DD/YYYY
        "%d/%m/%Y",  # DD/MM/YYYY
        "%d-%m-%Y",  # DD-MM-YYYY
        "%d.%m.%Y",  # DD.MM.YYYY
        "%Y%m%d",  # YYYYMD
        "%m%d%Y",  # MDYYYY
        "%d%m%Y"  # DMYYYY
    ]

    for date_format in date_formats:
        try:
            # 尝试将字符串转换为日期对象
            date_obj = datetime.strptime(date_str, date_format)
            # 返回标准的 YYYY-MM-DD 格式
            return True, date_obj.strftime("%Y-%m-%d")
        except ValueError:
            # 如果转换失败，继续尝试下一个格式
            continue

    # 如果所有格式都失败，返回 '0'
    return False, None


def is_date_within_range_and_days_left(target_date_str, start_date_str, end_date_str):
    """
    检查目标日期是否在指定的日期范围内，并计算到截止日期还剩下多少天。

    :param target_date_str: 目标日期字符串，格式为 'YYYY-MM-DD'
    :param start_date_str: 起始日期字符串，格式为 'YYYY-MM-DD'
    :param end_date_str: 结束日期字符串，格式为 'YYYY-MM-DD'
    :return: 一个元组 (是否在范围内, 剩余天数)
    """

    # 检查目标日期是否有效
    is_target_valid, target_date = is_valid_date(target_date_str)
    if not is_target_valid:
        return False, 0

    # 检查起始日期是否有效
    is_start_valid, start_date = is_valid_date(start_date_str)
    if not is_start_valid:
        return False, 0

    # 检查结束日期是否有效
    is_end_valid, end_date = is_valid_date(end_date_str)
    if not is_end_valid:
        return False, 0

    # 定义日期格式
    date_format = "%Y-%m-%d"  # 根据你的日期格式调整

    # 将字符串转换为 datetime 对象
    start_date = datetime.strptime(start_date, date_format)
    target_date = datetime.strptime(target_date, date_format)
    end_date = datetime.strptime(end_date, date_format)

    # 检查目标日期是否在指定的日期范围内
    is_within_range = start_date <= target_date <= end_date

    # 计算到截止日期还剩下多少天
    days_left = (end_date - target_date).days

    return is_within_range, days_left


def extract_and_format_date(url):
    """
    Extracts the date string from the input URL and converts it to standard date format (YYYY-MM-DD).

    Args:
        url (str): The input URL containing a date string in the format YYYYMMDDhhmmss

    Returns:
        str: The formatted date string in the format YYYY-MM-DD, or an empty string if no date is found.
    """
    # Define the regex pattern to match a 14-digit date format (YYYYMMDDhhmmss)
    date_pattern = re.compile(r'\d{14}')

    # Search for the pattern in the URL
    match = date_pattern.search(url)

    if match:
        date_str = match.group()  # Get the matched date string
        try:
            # Convert the date string to a datetime object
            date_obj = datetime.strptime(date_str, '%Y%m%d%H%M%S')
            # Format the datetime object to YYYY-MM-DD
            formatted_date = date_obj.strftime('%Y-%m-%d')
            return formatted_date
        except ValueError:
            # Handle possible errors in date conversion
            print(f"Error converting date: {date_str}")
            return None
    else:
        # Return an empty string if no date is found
        print("No date found in the URL.")
        return None

def get_weekday_number(date_str):
    # 将字符串转换为日期对象
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")

    # 获取星期几的数字，0 表示星期一，6 表示星期日
    weekday_number = date_obj.weekday()

    # 将0到6的数字转换为1到7的数字
    return weekday_number + 1


def save_send_report_info(args, project, branch, send_date):
    # 定义CSV文件的列名
    headers = ['project', 'branch', 'send_date']
    file_exists = os.path.exists(args.tt_message_info)

    updated = False
    rows = []

    if file_exists:
        # 读入现有的CSV文件
        with open(args.tt_message_info, mode='r', newline='') as file:
            reader = csv.DictReader(file)
            rows = list(reader)

        # 查找是否存在相同的项目并更新
        for row in rows:
            if row.get('project', '0') == project and row.get('branch', '0') == branch:
                print("\nBefore Update:", row)
                if send_date:
                    row['send_date'] = send_date
                if branch:
                    row['branch'] = branch
                print("After Update:", row)
                updated = True
                break

    if not updated:
        # 如果不存在相同的项目，则添加新的条目
        new_row = {
            'project': project,
            'branch': branch,
            'send_date': send_date
        }
        rows.append(new_row)
        print("New Entry:", new_row)

    # print("info:\n", rows)
    # 写入更新或新创建的内容到CSV文件
    with open(args.tt_message_info, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            try:
                # 过滤掉不在headers中的字段
                row_filtered = {key: format_as_text(row[key]) for key in headers if key in row}
                writer.writerow(row_filtered)
            except ValueError as e:
                print(f"Error writing row {row}: {e}")
                continue


def get_send_report_info(args, project, branch, delta):
    """
    读取 CSV 文件，查找与传递的 project 相同的行，并返回各列的内容。
    """
    file_path = args.tt_message_info
    remove_expired_data(args, file_path, 'send_date', delta)
    # 定义默认值
    default_values = {
        'project': '0',
        'branch': '0',
        'send_date': '0'
    }

    if not os.path.exists(file_path):
        print(f"file  {file_path} not exist")
        return default_values, False

    # 读取 CSV 文件
    with open(file_path, mode='r', newline='') as file:
        reader = csv.DictReader(file)
        rows = list(reader)

    # 查找是否存在相同的项目
    for row in rows:
        if row.get('project', '0') == project and row.get('branch', '0') == branch:
            return {
                       'project': row.get('project', '0'),
                       'branch': row.get('branch', '0'),
                       'send_date': row.get('send_date', '0')
                   }, True

    print(f"not find project {project} send config info in {file_path} use default")
    return default_values, False


def check_ack_data_list_route_v1(args):
    """
    Args:
        args:
    """
    local_date = args.current_date

    # 计算总行数
    with open(args.ack_all_info, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        total_rows = sum(1 for row in reader) - 1

    with open(args.ack_all_info, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        # for row in reader:
        for line_number, row in enumerate(reader, start=1):
            build_date = row.get('build_date', '0').strip("'")
            ignore_date_start = row.get('ignore_date_start', '0').strip("'")
            ignore_date_end = row.get('ignore_date_end', '0').strip("'")
            days_eol = int(row.get('days_eol', '0'))
            warning_delta = int(row.get('warning_delta', '0'))
            error_delta = int(row.get('error_delta', '0'))
            warning_duty = int(row.get('warning_duty', '0'))
            error_duty = int(row.get('error_duty', '0'))
            release_eol = row.get('release_eol', '0').strip("'")
            branch = row.get('branch', '0')
            project = row.get('project', '0')
            version = row.get('version', '0')
            area = row.get('area', '0')
            name = row.get('name', '0')

            valid_date = days_until_eol(str(local_date), str(build_date))
            error_branch = check_branch(branch)
            if valid_date < 0 or not error_branch:
                continue

            eol_days_left = days_until_eol(str(release_eol), str(local_date))

            start_error_date = adjust_date_by_days(release_eol, error_delta)
            warning_to_error_left = days_until_eol(str(start_error_date), str(local_date))
            is_within_range, ignore_days_left = is_date_within_range_and_days_left(
                str(local_date), ignore_date_start, ignore_date_end)

            # 判断是否需要发送警告消息
            need_error = (1 < days_eol < error_delta and 1 < eol_days_left < error_delta)
            need_warning = (error_delta < days_eol < warning_delta and error_delta < eol_days_left < warning_delta)
            is_within_critical_range = (is_within_range and ignore_days_left < 2) or 1 <= warning_to_error_left <= 2

            if ((need_error or need_warning) and not is_within_range) or is_within_critical_range:
                project_info, matched = get_send_report_info(args, project, branch, warning_duty)
                if not matched or is_within_critical_range:
                    print(f"is_within_critical_range force send: {is_within_critical_range}")
                    need_send_message = True
                else:
                    send_date = project_info.get('send_date', '0').strip("'")
                    send_date_days = days_until_eol(str(local_date), str(send_date))
                    print(f"previous_send_date: {send_date}")
                    print(f"already send days: {send_date_days}")
                    print(f"warning_duty: {warning_duty}")
                    print(f"error_duty: {error_duty}")
                    if need_warning and abs(send_date_days) >= warning_duty:
                        need_send_message = True
                    elif need_error and abs(send_date_days) >= error_duty:
                        need_send_message = True
                    else:
                        need_send_message = False
            else:
                need_send_message = False

            print(f"Line {line_number}/{total_rows} {local_date}\n")

            print(f"local_date:{local_date} build_date:{build_date} ignore_date_start:{ignore_date_start} "
                  f"ignore_date_end:{ignore_date_end} \ndays_eol:{days_eol} warning_delta:{warning_delta} "
                  f"error_delta:{error_delta} warning_duty:{warning_duty} error_duty:{error_duty} \n"
                  f"release_eol:{release_eol} "
                  f"valid_date:{valid_date}=({local_date})-({build_date})>=0  "
                  f"eol_days_left:{eol_days_left}=({release_eol})-({local_date}) \n"
                  f"start_error_date:{start_error_date}=({release_eol})-{error_delta} "
                  f"warning_to_error_left:{warning_to_error_left}=({start_error_date})-({local_date}) \n"
                  f"is_within_range:{is_within_range}=({ignore_date_start})<=({local_date})<=({ignore_date_end}) \n"
                  f"ignore_days_left:{ignore_days_left}=({local_date})-({ignore_date_end}) \n"
                  f"need_error:{need_error}=(1 < {days_eol} < {error_delta} and {eol_days_left} > 0) \n"
                  f"need_warning:{need_warning}=({error_delta} < {days_eol} < {warning_delta}"
                  f" and {eol_days_left} > 0)\n"
                  f"is_within_range {is_within_range}  is_within_critical_range:{is_within_critical_range} and"
                  f" {ignore_days_left} < 2  \n"
                  )
            """"""
            message_base = 'project:' + project + '\nname:' + name + '\nversion:' + version + \
                           '\nbranch:' + branch + '\narea:' + area
                           #'\nrelease_eol:' + release_eol + '\ndays_eol:' + row.get('days_eol', '0')
            if eol_days_left > 0:
                message_warning = '\n离项目过期时间' + release_eol + '还有约' + str(eol_days_left) + '天请及时安排升级'
            else:
                message_warning = ''

            if warning_to_error_left > 0:
                message_error = '\n离项目编译报错时间' + start_error_date + '还有约' + str(warning_to_error_left) + \
                                '天请及时安排升级'
            else:
                message_error = ''

            message_help = "\n更多帮助请阅读如下文档\n"
            message = message_base + message_error + message_warning + message_help

            cfg_lists = ["pm_id", "cmo_id", "grs_userid", "config_userid"]
            user_ids_lists = []

            for cfg_list in cfg_lists:
                user_ids = row.get(cfg_list, '0')
                user_ids = extract_and_split_valid(user_ids)
                for user_id in user_ids:
                    user_ids_lists.append(user_id)
            # 去重
            unique_user_ids_lists = list(set(user_ids_lists))
            # 排序
            sorted_unique_user_ids_lists = sorted(unique_user_ids_lists)
            users_str = ','.join(sorted_unique_user_ids_lists)

            # 综合判断是否需要发送消息
            if need_send_message:
                print(f"message: \n{message}\n")
                save_send_report_info(args, project, branch, str(local_date))
                send_user_message_cmd(args, users_str, message)



def check_ack_data_list_route(args):
    """
    Args:
        args:
    """
    local_date = args.current_date
    local_ack_all_info = os.path.join(args.out, "ack_all_info_final.csv")

    # 计算总行数
    with open(local_ack_all_info, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        total_rows = sum(1 for row in reader) - 1

    with open(local_ack_all_info, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        # for row in reader:
        for line_number, row in enumerate(reader, start=1):
            release_eol = row.get('release_eol', '0').strip("'")
            gerrit_projects = row.get('gerrit_projects', '0')
            gerrit_branch = row.get('gerrit_branch', '0')
            release_version = row.get('release_version', '0')
            message = row.get('message', '0')

            cfg_lists = ["pm_id", "cmo_id", "grs_userid", "config_userid"]
            user_ids_lists = []

            for cfg_list in cfg_lists:
                user_ids = row.get(cfg_list, '0')
                user_ids = extract_and_split_valid(user_ids)
                for user_id in user_ids:
                    user_ids_lists.append(user_id)
            # 去重
            unique_user_ids_lists = list(set(user_ids_lists))
            # 排序
            sorted_unique_user_ids_lists = sorted(unique_user_ids_lists)
            users_str = ','.join(sorted_unique_user_ids_lists)

            eol_days_left = days_until_eol(str(release_eol), str(local_date))
            weekday_number = get_weekday_number(str(local_date))
            print(f"Line {line_number}/{total_rows} date:{local_date} weekday:{weekday_number} left:{eol_days_left}\n")

            message_base =  message + '\n\n'  \
                            'project:' + gerrit_projects + '\nbranch:' + gerrit_branch + \
                           '\nversion:' + release_version + \
                           '\neol:' + release_eol + '\n' + \
                            'user:' + users_str+ '\n'
            if eol_days_left > 0:
                message_warning = '\n离版本过期时间' + release_eol + '还有约' + str(eol_days_left) + '天请及时安排升级'
            else:
                message_warning = ''

            message_help = "\n更多帮助请阅读如下文档\n"
            message = message_base + message_warning + message_help

            if 0 < eol_days_left < 90 and weekday_number ==1:
                need_send_message = True
            else:
                need_send_message = False
            # 综合判断是否需要发送消息
            if need_send_message:
                print(f"\n{message}\n")
                #save_send_report_info(args, project, branch, str(local_date))
                send_user_message_cmd(args, users_str, message)

def remove_expired_data(args, file_name, row_name, delta):
    """
    读取 CSV 文件，删除 `build_date` 列中的日期与当前日期相差 delta 天以上的行，并将更新后的信息写回原来的文件。
    """
    try:
        # 读取 CSV 文件内容
        with open(file_name, mode='r', newline='') as file:
            csv_reader = csv.DictReader(file)
            rows = list(csv_reader)

        # 获取当前日期
        today = datetime.now().date()

        # 过滤掉需要删除的行
        updated_rows = []
        for row in rows:
            build_date = row.get(row_name, '0')
            is_valid, build_date = is_valid_date(build_date)
            if is_valid:
                days_difference = days_until_eol(str(today), str(build_date))
                if days_difference < delta:
                    updated_rows.append(row)
                else:
                    print(f"out of date row {row}\nbuild date {build_date} current date {today} "
                          f"days_difference {days_difference} warning_delta {args.warning_delta}")

        # 写回更新后的信息
        with open(file_name, mode='w', newline='') as file:
            csv_writer = csv.DictWriter(file, fieldnames=csv_reader.fieldnames)
            csv_writer.writeheader()
            for row in updated_rows:
                try:
                    # 过滤掉不在headers中的字段
                    row_filtered = {key: format_as_text(row[key]) for key in csv_reader.fieldnames if key in row}
                    csv_writer.writerow(row_filtered)
                except ValueError as e:
                    print(f"Error writing row {row}: {e}")
                    continue

    except FileNotFoundError:
        print(f"File not found: {file_name}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def search_project_from_file(args, csv_file):
    """
    逐行读取 CSV 文件的内容，并打印每一行的内容及其行号和总行号。

    :param args: 其他参数（如果需要）
    :param csv_file: CSV 文件的路径
    """

    # 计算总行数
    with open(csv_file, mode='r', newline='') as file:
        total_rows = sum(1 for row in csv.reader(file))

    if os.path.exists(args.ack_all_info):
        os.remove(args.ack_all_info)
        print("this is full build,regenerate it")

    # 逐行读取并打印
    with open(csv_file, mode='r', newline='') as file:
        csv_reader = csv.reader(file)
        for line_number, row in enumerate(csv_reader, start=1):
            if row:  # 确保行不为空
                try:
                    url = row[0] if row else None  # 将列表转换成字符串
                    print(f"Line {line_number} of {total_rows}: {url}")
                    args.bootimg = url
                    android_common_kernel_build(args)
                except FileNotFoundError:
                    print(f"File not found: {csv_file}")
                except Exception as e:
                    print(f"An error occurred: {str(e)}")


def generate_ack_info(args):
    version_info = ""
    local_date = ""
    version = find_current_linux_version(args)
    if version:
        print("\nparse kernel version")
        version_info = parse_version(version)
        if version_info:
            print_version_info(version_info)
            tmp_date = extract_and_format_date(args.bootimg)
            if tmp_date:
                local_date = tmp_date
            else:
                local_date = args.current_date
            update_info = "\nLINUX_KERNEL_BUILD_DATE:" + local_date + "\n"
            print(update_info)
            update_kernel_info(update_info)

    return version, version_info, local_date

def android_common_kernel_build(args):
    build_info_dict = {}
    version, version_info, local_date = generate_ack_info(args)
    if version and version_info and local_date and args.type != 'build':
        build_info_dict = generate_ack_data_list_save(args, version, version_info, local_date)

    delete_list = ['boot.img', 'compile.ini', 'unpack', 'build.json' , 'merge_component_manifests.xml']
    delete_files_and_dirs(args.out, delete_list)
    write_info_to_file(args.kernelinfo)
    check_file_exists(args)
    if build_info_dict:
        remove_expired_data(args, args.ack_all_info, 'build_date', int(args.error_delta) / 2)
        check_build_status(args, build_info_dict)

def get_md5_hash_from_file(file_name, md5_sums_file):
    """
    Read the md5_sums.txt file and return the MD5 hash for the specified file name.

    :param file_name: The name of the file to look up
    :param md5_sums_file: Path to the md5_sums.txt file
    :return: MD5 hash of the specified file name, or None if not found
    """
    # 初始化一个字典来存储文件名和对应的 MD5 哈希值
    md5_dict = {}

    # 检查文件是否存在
    if not os.path.exists(md5_sums_file):
        print(f"File {md5_sums_file} does not exist.")
        return None

    # 读取 md5_sums.txt 文件
    with open(md5_sums_file, 'r') as md5_file:
        for line in md5_file:
            # 去除行末的换行符并按冒号分割
            parts = line.strip().split(':')
            if len(parts) >= 2:
                # 假设第一部分是文件名，第二部分是 MD5 哈希值
                file_name_in_line = parts[0].strip()
                md5sum = parts[1].strip()
                # 将文件名和 MD5 哈希值存储在字典中
                md5_dict[file_name_in_line] = md5sum

    # 根据传入的文件名返回对应的 MD5 哈希值
    return md5_dict.get(file_name, None)


def transfer_between_remote_local(args, transfer):
    ack_info_remote_url = "xxx/aosp-gki-image-local/ack_info/"
    download_configs = [
        {
            "remote_file_name": "ack_all_info.csv",
            "local_file_name": "ack_all_info.csv",
        },
        {
            "remote_file_name": "ack_all_info.xlsx",
            "local_file_name": "ack_all_info.xlsx",
        },
        {
            "remote_file_name": "ack_all_info_final.csv",
            "local_file_name": "ack_all_info_final.csv",
        },
        {
            "remote_file_name": "ack_all_info_final.xlsx",
            "local_file_name": "ack_all_info_final.xlsx",
        },
        {
            "remote_file_name": "final_paths.csv",
            "local_file_name": "final_paths.csv",
        },
        {
            "remote_file_name": "final_paths.xlsx",
            "local_file_name": "final_paths.xlsx",
        },
        {
            "remote_file_name": "ack_branch_info.xlsx",
            "local_file_name": "ack_branch_info.xlsx",
        },
        {
            "remote_file_name": "prj_cfg_info.csv",
            "local_file_name": "prj_cfg_info.csv",
        }
    ]
    if args.type == 'build':
        return

    md5_sums = os.path.join(args.out, "md5_sums.txt")
    if transfer == "download":
        if os.path.exists(md5_sums):
            os.remove(md5_sums)

    # 使用 for 循环遍历每个配置并下载文件
    for config in download_configs:
        remote_file_name = config["remote_file_name"]
        local_file_name = config["local_file_name"]
        remote_file = os.path.join(ack_info_remote_url, remote_file_name)
        local_file = os.path.join(args.out, local_file_name)

        if transfer == "download":
            common.download_file_from_remote(local_file, remote_file)
            file_names = [local_file_name]
            calculate_files_md5sum(file_names,md5_sums,args.out)
        else:
            md5sum_local = calculate_file_md5(local_file)
            md5sum_sever = get_md5_hash_from_file(local_file_name, md5_sums)
            print("{:<40} {} {}".format(local_file_name, md5sum_sever, md5sum_local))
            if md5sum_sever != md5sum_local:
                #print(f"MD5 hash for {md5sum_local}: {md5sum_sever}")
                common.upload_file(local_file, remote_file)
    return


def config_branch_from_file(args, file_path):
    if not os.path.exists(file_path):
        return
    try:
        # 加载Excel文件
        workbook = load_workbook(filename=file_path)

        # 假设数据在第一个工作表中
        sheet = workbook.active

        # 获取列标题
        headers = [cell.value for cell in sheet[1]]

        # 检查所需的列是否存在
        required_columns = ['gerrit_branch', 'config_userid']
        if not all(column in headers for column in required_columns):
            print(f"文件中缺少所需的列: {required_columns}")
            return

        # 获取所需列的索引
        gerrit_branch_index = headers.index('gerrit_branch')
        config_userid_index = headers.index('config_userid')

        # 打印列标题
        print(f"{'gerrit_branch':<20} {'config_userid':<20}")
        print("-" * 40)

        # 遍历每一行并打印所需列的内容
        for row in sheet.iter_rows(min_row=2, values_only=True):
            gerrit_branch = row[gerrit_branch_index]
            config_userid = row[config_userid_index]
            #print(f"{gerrit_branch:<20} {config_userid:<20}")
            config_userid = standardize_delimiters(config_userid)
            #print("current config_userid:", config_userid)
            save_config_project_info(args, gerrit_branch, config_userid)

    except FileNotFoundError:
        print(f"文件未找到: {file_path}")
    except Exception as e:
        print(f"发生错误: {e}")


def set_config_project_info(args):
    branches = extract_and_split_valid(args.branch)
    if branches:
        for branch in branches:
            print("current branch:", branch)
            config_userid = standardize_delimiters(args.config_userid)
            save_config_project_info(args, branch, config_userid)
    else:
        # 示例调用
        ack_branch_info = os.path.join(args.out, "ack_branch_info.xlsx")
        config_branch_from_file(args, ack_branch_info)


def serarch_project_from_server(args):
    ack_dump.get_all_project_info(args)
    output_csv = os.path.join(args.out, "final_paths.csv")
    output_xlsx = os.path.join(args.out, "final_paths.xlsx")
    search_project_from_file(args, output_csv)
    csv_file = os.path.join(args.out, "ack_all_info.csv")
    xlsx_file = os.path.join(args.out, "ack_all_info.xlsx")
    optimized_csv_file = os.path.join(args.out, "ack_all_info_final.csv")
    optimized_xlsx_file = os.path.join(args.out, "ack_all_info_final.xlsx")
    ack_info.ack_optimize_and_save_csv(args, csv_file, optimized_csv_file)
    ack_info.ack_csv_to_xlsx(args, output_csv, output_xlsx)
    ack_info.ack_csv_to_xlsx(args, csv_file, xlsx_file)
    ack_info.ack_csv_to_xlsx(args, optimized_csv_file, optimized_xlsx_file)


def main(args):

    transfer_between_remote_local(args, "download")

    if args.type == 'send':
        serarch_project_from_server(args)
        check_ack_data_list_route(args)
    elif args.type == 'config':
        set_config_project_info(args)
    elif args.type == 'dump':
        serarch_project_from_server(args)
        check_ack_data_list_route(args)
    else:
        # print("ack build")
        android_common_kernel_build(args)

    transfer_between_remote_local(args, "upload")


if __name__ == "__main__":
    args = parse_cmd_args()
    try:
        start_time = datetime.now()
        print("Begin time:", start_time.strftime("%Y-%m-%d %H:%M:%S"))
        main(args)
        check_file_exists(args)
        end_time = datetime.now()
        print("End time:", end_time.strftime("%Y-%m-%d %H:%M:%S"))
        elapsed_time = end_time - start_time
        print("Total time:", common.format_duration(elapsed_time))
    except SystemExit:
        check_file_exists(args)
        sys.exit(1)
    except Exception as e:
        check_file_exists(args)
        print("error: {}".format(e))

